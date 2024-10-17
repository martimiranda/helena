from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QLabel, QLineEdit, QTableWidget, 
    QTableWidgetItem, QPushButton, QWidget, QVBoxLayout, QGridLayout, QSizePolicy, 
    QHeaderView, QScrollBar,QCheckBox, QMessageBox, QDialog, QDateEdit, QFrame, QFileDialog
)
from PyQt6.QtGui import QPixmap, QFont, QIcon, QPainter
from PyQt6.QtCore import Qt,QDate
from PyQt6.QtSvgWidgets import QSvgWidget
from PyQt6.QtSvg import QSvgRenderer
from datetime import datetime
import sqlite3, sys, re, shutil
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font,  PatternFill


class DialogoConfirmacion(QDialog):
    def __init__(self, parent=None, mensajeClase= None):
        super().__init__(parent)


        self.setWindowTitle('Confirmar Eliminació')
        self.setFixedSize(600, 400)  # Tamaño fijo

        # Crear diseño vertical
        layout = QGridLayout()
        layout_arriba = QGridLayout()
        layout_abajo = QGridLayout()

        # Etiqueta con el mensaje
        icono = QLabel()
        svg_renderer = QSvgRenderer('assets/error-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        icono.setPixmap(svg_pixmap)
        icono.setAlignment(Qt.AlignmentFlag.AlignCenter)

        mensaje = QLabel("¿Vols eliminar per sempre aquest client?\nSi l'elimines no es podra recuperar.")
        mensaje.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
            }
        """)

        if mensajeClase is not None:
            mensaje = QLabel("¿Vols eliminar per sempre aquest servei?\nSi l'elimines no es podra recuperar.")
            mensaje.setStyleSheet("""
                QLabel {
                    font-size: 16px;
                    font-weight: bold;
                }
            """)

    

        layout_arriba.addWidget(icono,0,0)
        layout_arriba.addWidget(mensaje,0,1)
        layout_arriba.setColumnStretch(0,5)
        layout_arriba.setColumnStretch(1,5)



        # Etiqueta con el ícono
       

        # Botones de confirmación
        aceptar_btn = QPushButton('Acceptar')
        aceptar_btn.setStyleSheet(""" QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
        }
        
        """)
        aceptar_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        aceptar_btn.clicked.connect(self.aceptar)

        cancelar_btn = QPushButton('Cancelar')
        cancelar_btn.setStyleSheet(""" QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
        }
        
        """)
        cancelar_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        cancelar_btn.clicked.connect(self.rechazar)

        layout_abajo.addWidget(aceptar_btn,0,0)
        layout_abajo.addWidget(cancelar_btn,0,1)
        layout_abajo.setColumnStretch(0,5)
        layout_abajo.setColumnStretch(1,5)

        layout.addLayout(layout_arriba,0,0)
        layout.addLayout(layout_abajo,1,0)
        layout.setRowStretch(0,8)
        layout.setRowStretch(1,2)


        self.setLayout(layout)

    def aceptar(self):
        self.accept()  # Aceptar el diálogo

    def rechazar(self):
        self.reject()  # Rechazar el diálogo


class VentanaError(QDialog):
    def __init__(self, mensaje, parent=None):
        super().__init__(parent)
        
        # Configuración de la ventana de error
        self.setWindowTitle("Error")
        self.setFixedSize(600, 400)
        # Crear widgets
        label = QLabel(mensaje)
        label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
            }
        """)
        botonCerrar = QPushButton("Acceptar")

        botonCerrar.setStyleSheet(""" QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
        }
        
        """)
        botonCerrar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)



        svg_renderer = QSvgRenderer('assets/error-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        labelIcono = QLabel()
        labelIcono.setPixmap(svg_pixmap)
        labelIcono.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Layout
        layoutCentral = QGridLayout()
        layoutCentral.addWidget(labelIcono,0,0)
        layoutCentral.addWidget(label,0,1)
        layoutCentral.setColumnStretch(0,5)
        layoutCentral.setColumnStretch(1,5)

        layoutBoton = QGridLayout()
        layoutBoton.addWidget(QLabel(),0,0)
        layoutBoton.addWidget(botonCerrar,0,1)
        layoutBoton.addWidget(QLabel(),0,2)
        layoutBoton.setColumnStretch(0,3)
        layoutBoton.setColumnStretch(1,3)
        layoutBoton.setColumnStretch(2,3)




        layout = QGridLayout()
        layout.addLayout(layoutCentral,0,0)
        layout.addLayout(layoutBoton,1,0)
        layout.setRowStretch(0,8)
        layout.setRowStretch(1,2)

        self.setLayout(layout)
        
        # Conectar el botón de cerrar
        botonCerrar.clicked.connect(self.close)

class VentanaServicios(QMainWindow):
    def __init__(self,parent= None):
        super().__init__(parent)

        # Guardar la variable recibida
        self.parent_window = parent
        

        self.id_cliente = self.parent_window.id_cliente
        self.nombre = self.parent_window.nom_value_label.text()
        print(self.id_cliente)

        # Configurar la ventana
        self.setWindowTitle("Serveis de "+ self.nombre)

        self.generar_contenido()
        self.show()



    def generar_contenido(self):
        self.setStyleSheet("background-color: white;")
        self.imagen_fondo_label = QLabel(self)
        pixmap = QPixmap('assets/icono.png')  # Cargar la imagen
        self.imagen_fondo_label.setPixmap(pixmap)
        self.imagen_fondo_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.estructuraExterna()

    def estructuraExterna(self):
        
        self.contenido_central = QGridLayout()
        self.layoutFondo = QGridLayout()
       
        self.layoutFondo.addWidget(QLabel(), 0, 0)
        self.layoutFondo.addWidget(self.imagen_fondo_label,0,1)
        self.layoutFondo.addLayout(self.contenido_central, 0, 1)
        self.layoutFondo.addWidget(QLabel(), 0, 2)
        self.layoutFondo.setColumnStretch(0, 1)
        self.layoutFondo.setColumnStretch(1, 8)
        self.layoutFondo.setColumnStretch(2, 1)

        contenedor_widget = QWidget()
        contenedor_widget.setLayout(self.layoutFondo)
        self.setCentralWidget(contenedor_widget)
        self.estructura_central()

    def estructura_central(self):
        self.cargarEstructuraSuperior()
        self.cargarCreacionInferior()
        self.cargarEdicionInferior()
        self.contenido_central.addWidget(self.label_superior, 0, 0)
        self.contenido_central.addWidget(self.tablaClientes, 1, 0)
        self.contenido_central.addWidget(self.no_data_client, 1, 0)
        self.contenido_central.addWidget(self.estructuraCreacionInferior,2,0)
        self.contenido_central.addWidget(self.estructuraEdicionInferior,2,0)


        self.contenido_central.setRowStretch(0, 1)
        self.contenido_central.setRowStretch(1, 4)
        self.contenido_central.setRowStretch(2, 5)


    def cargarEstructuraSuperior(self):
        self.label_superior = QLabel("Serveis de "+ self.nombre)
        self.label_superior.setStyleSheet("""
            font-weight: 100; /* Peso de la fuente (más fino) */
            background-color: rgba(255, 165, 0, 100);
            border: 1px solid #d3d3d3;
            border-radius: 10px;
            font-weight: bold;  /* Encabezados en negrita */
            font-size: 18px;
            text-align: center;

        """)
        self.label_superior.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tablaClientes = QTableWidget(self)
        self.tablaClientes.setColumnCount(4)  # Número de columnas
        self.tablaClientes.setHorizontalHeaderLabels(["Data", "Servei", "Preu"])  # Nombres de las columnas
        self.tablaClientes.setRowCount(0)  # Inicialmente sin filas
        self.configure_table()

        # Conectar el evento de scroll
        self.tablaClientes.verticalScrollBar().valueChanged.connect(self.load_more_serveis)
        self.tablaClientes.cellClicked.connect(self.on_servicio_click)

        self.tablaClientes.hideColumn(3)
        # Contadores para la paginación
        self.offset = 0
        self.limit = 30

        self.no_data_client = QPushButton(self)
        self.no_data_client.setText('Aquest client no ha fet cap servei')
        self.no_data_client.setStyleSheet("""
            QPushButton {
                font-weight: 100;
                background-color: rgba(211, 211, 211, 0.4);                    
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                font-size: 18px;
                color: black;
                font-weight: bold;
            }
        """)
        svg_renderer = QSvgRenderer('assets/information-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        


        self.no_data_client.setIcon(icon)
        self.no_data_client.setIconSize(svg_pixmap.size())
        self.no_data_client.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.no_data_client.hide()

        self.cargar_servicios()

    def configure_table(self):
        header = self.tablaClientes.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        # Ocultar el encabezado vertical
        self.tablaClientes.verticalHeader().setVisible(False)

        # Estilo para la tabla
        self.tablaClientes.setStyleSheet("""
            QTableWidget {
                background: transparent;
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                
            }
            QHeaderView::section {
                background-color: rgba(255, 165, 0, 100);
                border: none;
                font-weight: bold;
                font-size: 18px;
                border-radius: 10px;
            }
            QTableWidget::item {
                background-color: rgba(255, 255, 255, 150);
                border: 1px solid #d3d3d3;
            }

            QTableWidget::item:selected {
                background-color: rgba(255, 255, 255, 150); /* Mantener el fondo igual */
                color: black;  /* El color del texto sigue siendo negro (o cualquier color que prefieras) */
            }
            
        """)
    
    def cargar_servicios(self):
        self.offset = 0
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()
        try:
            self.query = '''
                    SELECT Fecha, Servei, Preu, Id
                    FROM Serveis
                    WHERE Client_id = ?
                    ORDER BY Fecha DESC
                    LIMIT ? OFFSET ?;
                '''

            params = (f'{self.id_cliente}', f'{self.limit}',f'{self.offset}')

            self.params_list = list(params)
            cursor.execute(self.query, params)

            rows = cursor.fetchall()

            if len(rows) == 0:
                # Si no hay resultados, ocultar la tabla y mostrar un mensaje
                self.tablaClientes.hide()
                self.no_data_client.show()

            else:
                # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                self.tablaClientes.setRowCount(0)
                for row in rows:
                    row_position = self.tablaClientes.rowCount()
                    self.tablaClientes.insertRow(row_position)
                    for column, value in enumerate(row):
                        if column == 2:  # Columna "Preu"
                            # Si el valor es un número con decimal .0, eliminarlo
                            if isinstance(value, float) and value.is_integer():
                                value = int(value)
                            # Convertir a cadena y agregar el símbolo de euro
                            value = f'{value}€'
                        elif column == 0:
                            value = self.fecha_a_string(value)
                        item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                        item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                        self.tablaClientes.setItem(row_position, column, item)

                self.no_data_client.hide()
                self.tablaClientes.show()
                

        finally:
            # Asegurarse de cerrar la conexión
            conn.close()

    def load_more_serveis(self):
        if self.tablaClientes.verticalScrollBar().value() == self.tablaClientes.verticalScrollBar().maximum():

            self.offset += self.limit
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            self.params_list[-1] = f'{self.offset}'
            params = tuple(self.params_list)
            try:
                cursor.execute(self.query, params)

                rows = cursor.fetchall()

                if len(rows) == 0:
                    pass
                else:
                    # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                    for row in rows:
                        row_position = self.tablaClientes.rowCount()
                        self.tablaClientes.insertRow(row_position)
                        for column, value in enumerate(row):
                            if column == 2:  # Columna "Preu"
                                # Si el valor es un número con decimal .0, eliminarlo
                                if isinstance(value, float) and value.is_integer():
                                    value = int(value)
                                # Convertir a cadena y agregar el símbolo de euro
                                value = f'{value}€'
                            elif column == 0:
                                 value = self.fecha_a_string(value)
                            item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                            item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                            self.tablaClientes.setItem(row_position, column, item)

                
            finally:
                # Asegurarse de cerrar la conexión
                conn.close()

    def string_a_fecha(self,fecha_str):
        return datetime.strptime(fecha_str, '%d/%m/%Y').date()

    # Conversión de objeto date a string
    def fecha_a_string(self,fecha_str):
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')  # Convierte la cadena a objeto datetime
        return fecha_obj.strftime('%d/%m/%Y')

    def cargarCreacionInferior(self):
        self.estructuraCreacionInferior = QWidget(self)  
        # Estilo aplicado al widget que contiene el layout
        self.estructuraCreacionInferior.setStyleSheet("""
            background-color: white;
            border: 1px solid #d3d3d3;
        """)
        
        estilosInput= """
            QLineEdit {
                border: 2px solid orange;
                border-radius: 15px;  /* Cambia este valor para ajustar el redondeado */
                padding: 5px;  /* Añade un poco de espacio interior */
                font-size: 16px;  /* Tamaño de fuente */
            }
        """

        label_style = """
            font-size: 16px;
            font-weight: bold;
            border: none;
        """

        estiloBoton= """ QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
        }
        
        """

        # Crear el layout principal de la ficha de usuario
        layout_fichaCreacionServicio = QGridLayout(self.estructuraCreacionInferior)

        label = QLabel("Creació de servei")
        label.setStyleSheet("""
            font-weight: 100; /* Peso de la fuente (más fino) */
            background-color: rgba(255, 165, 0, 100);
            border: 1px solid #d3d3d3;
            border-radius: 10px;
            font-weight: bold;  /* Encabezados en negrita */
            font-size: 18px;
            text-align: center;

        """)

        layout_inputs = QGridLayout()

        fecha_label = QLabel("Data:")
        fecha_label.setStyleSheet(label_style)
        fecha_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        servicio_label = QLabel("Servei:")
        servicio_label.setStyleSheet(label_style)
        servicio_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        precio_label = QLabel("Preu:")
        precio_label.setStyleSheet(label_style)
        precio_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.inputFechaCreacion = QLineEdit(self)
        try:
            fecha_hoy = QDate.currentDate().toString('dd/MM/yyyy')
            if not fecha_hoy:
                raise ValueError("Fecha no válida")
        except Exception as e:
            # Si hay algún error al obtener la fecha, mostrar un texto predeterminado
            fecha_hoy = "dd/mm/yyyy"
        self.inputFechaCreacion.setText(fecha_hoy)
        self.inputFechaCreacion.setStyleSheet(estilosInput)
        self.inputFechaCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


        self.inputServeiCreacion = QLineEdit(self)
        self.inputServeiCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputServeiCreacion.setStyleSheet(estilosInput)

        self.inputPreuCreacion = QLineEdit(self)
        self.inputPreuCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputPreuCreacion.setStyleSheet(estilosInput)

        layout_inputs.addWidget(fecha_label,0,0)
        layout_inputs.addWidget(servicio_label,1,0)
        layout_inputs.addWidget(precio_label,2,0)
        layout_inputs.addWidget(self.inputFechaCreacion,0,1)
        layout_inputs.addWidget(self.inputServeiCreacion,1,1)
        layout_inputs.addWidget(self.inputPreuCreacion,2,1)
        layout_inputs.setColumnStretch(0,5)
        layout_inputs.setColumnStretch(1,5)


        create_button = QPushButton(self)
        create_button.setText('Afegir nou servei')
        create_button.setStyleSheet(estiloBoton)
        create_button.clicked.connect(self.crear_servicio)


        svg_renderer = QSvgRenderer('assets/crear-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        iconCrear = QIcon(svg_pixmap)

        create_button.setIcon(iconCrear)
        create_button.setIconSize(svg_pixmap.size())
        create_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        close_button = QPushButton(self)
        close_button.setText('Tancar finestra')
        close_button.setStyleSheet(estiloBoton)
        close_button.clicked.connect(self.close)


        svg_renderer = QSvgRenderer('assets/cross-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        iconClose = QIcon(svg_pixmap)

        close_button.setIcon(iconClose)
        close_button.setIconSize(svg_pixmap.size())
        close_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        layoutBotones = QGridLayout()
        layoutBotones.addWidget(create_button,0,0)
        layoutBotones.addWidget(close_button,0,1)
        layoutBotones.setColumnStretch(0,5)
        layoutBotones.setColumnStretch(1,5)


        layout_fichaCreacionServicio.addWidget(label,0,0)
        layout_fichaCreacionServicio.addLayout(layout_inputs,1,0)
        layout_fichaCreacionServicio.addLayout(layoutBotones,2,0)
        layout_fichaCreacionServicio.setRowStretch(0,1)
        layout_fichaCreacionServicio.setRowStretch(1,7)
        layout_fichaCreacionServicio.setRowStretch(2,2)




    def crear_servicio(self):
        fecha = self.inputFechaCreacion.text().strip()
        servei = self.inputServeiCreacion.text()
        preu = self.inputPreuCreacion.text().strip()
        preu = preu.replace(",", ".")

        patron = r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$"
    
        # Comprobar si la fecha coincide con el patrón
        if not re.match(patron, fecha):
            ventana_error = VentanaError("La data no te un format valid", self)
            ventana_error.exec()
            return
        fecha = self.string_a_fecha(fecha)
            
        if len(servei.strip()) == 0:
            ventana_error = VentanaError("El servei no pot estar buit", self)
            ventana_error.exec()
            return
        try:
            # Intentar convertir el valor a float
            preu = float(preu)
        except ValueError:
            # Si no es un número válido, devuelve False o un mensaje de error
            ventana_error = VentanaError("El preu només accepta numeros i no pot estar buit", self)
            ventana_error.exec()
            return

        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            
            # Query de inserción para la tabla Serveis
            insert_query = '''
                INSERT INTO Serveis (Client_id, Fecha, Preu, Servei)
                VALUES (?, ?, ?, ?);
            '''
            
           
            
            # Ejecutar la query con los parámetros
            cursor.execute(insert_query, (self.id_cliente, fecha, preu, servei))
            
            # Confirmar los cambios
            conn.commit()
            
            self.cargar_servicios()
            self.inputServeiCreacion.setText("")
            self.inputPreuCreacion.setText("")
            
        except sqlite3.Error as e:
            print(f"Error al insertar el servicio: {e}")
            
        finally:
            # Cerrar la conexión
            conn.close()

        
        
        
         
            

        













    
    def cargarEdicionInferior(self):
        self.estructuraEdicionInferior = QWidget(self)  
        # Estilo aplicado al widget que contiene el layout
        self.estructuraEdicionInferior.setStyleSheet("""
            background-color: white;
            border: 1px solid #d3d3d3;
        """)
        self.estructuraEdicionInferior.hide()
        
        estilosInput= """
            QLineEdit {
                border: 2px solid orange;
                border-radius: 15px;  /* Cambia este valor para ajustar el redondeado */
                padding: 5px;  /* Añade un poco de espacio interior */
                font-size: 16px;  /* Tamaño de fuente */
            }
        """

        label_style = """
            font-size: 16px;
            font-weight: bold;
            border: none;
        """

        estiloBoton= """ QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
        }
        
        """

        svg_renderer_edit = QSvgRenderer('assets/edit-svg.svg')
        svg_pixmap_edit = QPixmap(24, 24)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap_edit.fill(Qt.GlobalColor.transparent)
        painter_edit = QPainter(svg_pixmap_edit)
        svg_renderer_edit.render(painter_edit)
        painter_edit.end()

        icon_edit = QIcon(svg_pixmap_edit)

        svg_renderer = QSvgRenderer('assets/tick-svg.svg')
        svg_pixmap = QPixmap(24, 24)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()

        icon = QIcon(svg_pixmap)

        svg_renderer_drop = QSvgRenderer('assets/trash-svg.svg')
        svg_pixmap_drop = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap_drop.fill(Qt.GlobalColor.transparent)
        painter_drop = QPainter(svg_pixmap_drop)
        svg_renderer_drop.render(painter_drop)
        painter_drop.end()

        icon_drop = QIcon(svg_pixmap_drop)


        # Crear el layout principal de la ficha de usuari

        # Crear el layout principal de la ficha de usuario
        layout_fichaEdicionServicio = QGridLayout(self.estructuraEdicionInferior)

        label = QLabel("Edició de servei")
        label.setStyleSheet("""
            font-weight: 100; /* Peso de la fuente (más fino) */
            background-color: rgba(255, 165, 0, 100);
            border: 1px solid #d3d3d3;
            border-radius: 10px;
            font-weight: bold;  /* Encabezados en negrita */
            font-size: 18px;
            text-align: center;

        """)

        layout_inputs = QGridLayout()

        fecha_label = QLabel("Data:")
        fecha_label.setStyleSheet(label_style)
        fecha_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        servicio_label = QLabel("Servei:")
        servicio_label.setStyleSheet(label_style)
        servicio_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        precio_label = QLabel("Preu:")
        precio_label.setStyleSheet(label_style)
        precio_label.setAlignment(Qt.AlignmentFlag.AlignCenter)


        self.fecha_label_value = QLabel()
        self.fecha_label_value.setStyleSheet(label_style)
        self.fecha_label_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.servicio_label_value = QLabel()
        self.servicio_label_value.setStyleSheet(label_style)
        self.servicio_label_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.precio_label_value = QLabel()
        self.precio_label_value.setStyleSheet(label_style)
        self.precio_label_value.setAlignment(Qt.AlignmentFlag.AlignCenter)



        self.inputFechaEdicion = QLineEdit(self)
        self.inputFechaEdicion.setStyleSheet(estilosInput)
        self.inputFechaEdicion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputFechaEdicion.returnPressed.connect(self.aceptar_fecha)
        self.inputFechaEdicion.hide()


        self.inputServeiEdicion = QLineEdit(self)
        self.inputServeiEdicion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputServeiEdicion.setStyleSheet(estilosInput)
        self.inputServeiEdicion.returnPressed.connect(self.aceptar_servei)
        self.inputServeiEdicion.hide()

        self.inputPreuEdicion = QLineEdit(self)
        self.inputPreuEdicion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputPreuEdicion.setStyleSheet(estilosInput)
        self.inputPreuEdicion.returnPressed.connect(self.aceptar_preu)
        self.inputPreuEdicion.hide()

        self.edit_fecha_button = QPushButton(self)
        self.edit_fecha_button.setText('Editar')
        self.edit_fecha_button.setStyleSheet(estiloBoton)
        self.edit_fecha_button.setIcon(icon_edit)
        self.edit_fecha_button.setIconSize(svg_pixmap_edit.size())
        self.edit_fecha_button.clicked.connect(self.editar_fecha)
        self.edit_fecha_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.edit_servei_button = QPushButton(self)
        self.edit_servei_button.setText('Editar')
        self.edit_servei_button.setStyleSheet(estiloBoton)
        self.edit_servei_button.setIcon(icon_edit)
        self.edit_servei_button.setIconSize(svg_pixmap_edit.size())
        self.edit_servei_button.clicked.connect(self.editar_servei)
        self.edit_servei_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.edit_preu_button = QPushButton(self)
        self.edit_preu_button.setText('Editar')
        self.edit_preu_button.setStyleSheet(estiloBoton)
        self.edit_preu_button.setIcon(icon_edit)
        self.edit_preu_button.setIconSize(svg_pixmap_edit.size())
        self.edit_preu_button.clicked.connect(self.editar_preu)
        self.edit_preu_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.accept_fecha_button = QPushButton(self)
        self.accept_fecha_button.setText('Acceptar')
        self.accept_fecha_button.setStyleSheet(estiloBoton)
        self.accept_fecha_button.setIcon(icon)
        self.accept_fecha_button.setIconSize(svg_pixmap.size())
        self.accept_fecha_button.clicked.connect(self.aceptar_fecha)
        self.accept_fecha_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_fecha_button.hide()

        self.accept_servei_button = QPushButton(self)
        self.accept_servei_button.setText('Acceptar')
        self.accept_servei_button.setStyleSheet(estiloBoton)
        self.accept_servei_button.setIcon(icon)
        self.accept_servei_button.setIconSize(svg_pixmap.size())
        self.accept_servei_button.clicked.connect(self.aceptar_servei)
        self.accept_servei_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_servei_button.hide()

        self.accept_preu_button = QPushButton(self)
        self.accept_preu_button.setText('Acceptar')
        self.accept_preu_button.setStyleSheet(estiloBoton)
        self.accept_preu_button.setIcon(icon)
        self.accept_preu_button.setIconSize(svg_pixmap.size())
        self.accept_preu_button.clicked.connect(self.aceptar_preu)
        self.accept_preu_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_preu_button.hide()

        layout_inputs.addWidget(fecha_label,0,0)
        layout_inputs.addWidget(servicio_label,1,0)
        layout_inputs.addWidget(precio_label,2,0)
        layout_inputs.addWidget(self.inputFechaEdicion,0,1)
        layout_inputs.addWidget(self.inputServeiEdicion,1,1)
        layout_inputs.addWidget(self.inputPreuEdicion,2,1)
        layout_inputs.addWidget(self.fecha_label_value,0,1)
        layout_inputs.addWidget(self.servicio_label_value,1,1)
        layout_inputs.addWidget(self.precio_label_value,2,1)
        layout_inputs.addWidget(self.edit_fecha_button,0,2)
        layout_inputs.addWidget(self.edit_servei_button,1,2)
        layout_inputs.addWidget(self.edit_preu_button,2,2)
        layout_inputs.addWidget(self.accept_fecha_button,0,2)
        layout_inputs.addWidget(self.accept_servei_button,1,2)
        layout_inputs.addWidget(self.accept_preu_button,2,2)
        layout_inputs.setColumnStretch(0,3)
        layout_inputs.setColumnStretch(1,3)
        layout_inputs.setColumnStretch(2,3)



        create_button = QPushButton(self)
        create_button.setText('Afegir nou servei')
        create_button.setStyleSheet(estiloBoton)
        create_button.clicked.connect(self.mostrar_creacion)


        svg_renderer = QSvgRenderer('assets/crear-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        iconCrear = QIcon(svg_pixmap)

        create_button.setIcon(iconCrear)
        create_button.setIconSize(svg_pixmap.size())
        create_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        close_button = QPushButton(self)
        close_button.setText('Tancar finestra')
        close_button.setStyleSheet(estiloBoton)
        close_button.clicked.connect(self.close)


        svg_renderer = QSvgRenderer('assets/cross-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        iconClose = QIcon(svg_pixmap)

        close_button.setIcon(iconClose)
        close_button.setIconSize(svg_pixmap.size())
        close_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        drop_button = QPushButton(self)
        drop_button.setText('Eliminar servei')
        drop_button.clicked.connect(self.eliminar_servei)
        drop_button.setIcon(icon_drop)
        drop_button.setIconSize(svg_pixmap_drop.size())
        drop_button.setStyleSheet(estiloBoton) 
        drop_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        layoutBotones = QGridLayout()
        layoutBotones.addWidget(create_button,0,0)
        layoutBotones.addWidget(close_button,0,1)
        layoutBotones.addWidget(drop_button,0,2)
        layoutBotones.setColumnStretch(0,3)
        layoutBotones.setColumnStretch(1,3)
        layoutBotones.setColumnStretch(2,3)



        layout_fichaEdicionServicio.addWidget(label,0,0)
        layout_fichaEdicionServicio.addLayout(layout_inputs,1,0)
        layout_fichaEdicionServicio.addLayout(layoutBotones,2,0)
        layout_fichaEdicionServicio.setRowStretch(0,1)
        layout_fichaEdicionServicio.setRowStretch(1,7)
        layout_fichaEdicionServicio.setRowStretch(2,2)

    def mostrar_creacion(self):
        self.estructuraEdicionInferior.hide()
        self.estructuraCreacionInferior.show()

    def on_servicio_click(self, row, column):
        self.id_servicio = self.tablaClientes.item(row, 3).text()
        fecha = self.tablaClientes.item(row, 0).text()
        servei = self.tablaClientes.item(row, 1).text()
        preu = self.tablaClientes.item(row, 2).text()
        self.fecha_label_value.setText(fecha)
        self.servicio_label_value.setText(servei)
        self.precio_label_value.setText(preu)
        self.hide_edit_servei()
        self.estructuraCreacionInferior.hide()
        self.estructuraEdicionInferior.show()

    def editar_servei(self):
        self.servicio_label_value.hide()
        self.edit_servei_button.hide()
        self.inputServeiEdicion.setText(self.servicio_label_value.text())
        self.inputServeiEdicion.show()
        self.accept_servei_button.show()

    def editar_fecha(self):
        self.fecha_label_value.hide()
        self.edit_fecha_button.hide()
        self.inputFechaEdicion.setText(self.fecha_label_value.text())
        self.inputFechaEdicion.show()
        self.accept_fecha_button.show()
    
    def editar_preu(self):
        self.precio_label_value.hide()
        self.edit_preu_button.hide()
        texto_editado = self.precio_label_value.text()[:-1]
        texto_editado = texto_editado.replace('.', ',')
        self.inputPreuEdicion.setText(texto_editado)
        self.inputPreuEdicion.show()
        self.accept_preu_button.show()


    def aceptar_servei(self):
        servei = self.inputServeiEdicion.text()
        if len(servei.strip()) == 0:
            ventana_error = VentanaError("El servei no pot estar buit", self)
            ventana_error.exec()
            return

        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()

        try:
            
            update_query = '''
                UPDATE Serveis
                SET Servei = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(update_query, (servei, self.id_servicio))
            
            # Confirmar los cambios
            conn.commit()

            self.servicio_label_value.setText(servei)

            self.cargar_servicios()

            

            self.edit_servei_button.show()
            self.servicio_label_value.show()
            self.inputServeiEdicion.hide()
            self.accept_servei_button.hide()



            
        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")
            
        finally:
            conn.close()
        
        

    def aceptar_fecha(self):
        fecha = self.inputFechaEdicion.text().strip()
        patron = r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$"
        fecha_str = fecha
    
        if not re.match(patron, fecha):
            ventana_error = VentanaError("La data no te un format valid", self)
            ventana_error.exec()
            return
        fecha = self.string_a_fecha(fecha)

        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()            
        try:
            
            update_query = '''
                UPDATE Serveis
                SET Fecha = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(update_query, (fecha, self.id_servicio))
            
            # Confirmar los cambios
            conn.commit()

            self.fecha_label_value.setText(fecha_str)

            self.cargar_servicios()

            

            self.edit_fecha_button.show()
            self.fecha_label_value.show()
            self.inputFechaEdicion.hide()
            self.accept_fecha_button.hide()



            
        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")
            
        finally:
            conn.close()
    
    def aceptar_preu(self):
        preu = self.inputPreuEdicion.text().strip()
        preu = preu.replace(",", ".")
        preuMostrar = self.inputPreuEdicion.text().strip()
        preuMostrar = f'{preuMostrar}€'

        try:
            # Intentar convertir el valor a float
            preu = float(preu)
        except ValueError:
            # Si no es un número válido, devuelve False o un mensaje de error
            ventana_error = VentanaError("El preu només accepta numeros i no pot estar buit", self)
            ventana_error.exec()
            return
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()            
        try:
            
            update_query = '''
                UPDATE Serveis
                SET Preu = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(update_query, (preu, self.id_servicio))
            
            # Confirmar los cambios
            conn.commit()

            self.precio_label_value.setText(preuMostrar)

            self.cargar_servicios()


            self.edit_preu_button.show()
            self.precio_label_value.show()
            self.inputPreuEdicion.hide()
            self.accept_preu_button.hide()



            
        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")
            
        finally:
            conn.close()
        
    
    def eliminar_servei(self):
        dialogo = DialogoConfirmacion(self, "servei")
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            try:
                conn = sqlite3.connect('data/data.db')  # Conectar a la base de datos
                cursor = conn.cursor()

                # Supongamos que `self.id_cliente` contiene el ID del cliente a eliminar
                delete_query = 'DELETE FROM Serveis WHERE Id = ?'
                cursor.execute(delete_query, (self.id_servicio,))

                # Confirmar cambios en la base de datos
                conn.commit()

                self.cargar_servicios()

                self.mostrar_creacion()



            except sqlite3.Error as e:
                print(f"Error al eliminar el cliente: {e}")

            finally:
                conn.close()

    def hide_edit_servei(self):
        self.fecha_label_value.show()
        self.inputFechaEdicion.hide()
        self.edit_fecha_button.show()
        self.accept_fecha_button.hide()
        self.servicio_label_value.show()
        self.inputServeiEdicion.hide()
        self.edit_servei_button.show()
        self.accept_servei_button.hide()
        self.precio_label_value.show()
        self.inputPreuEdicion.hide()
        self.edit_preu_button.show()
        self.accept_preu_button.hide()




    































class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.inicializar_ui()
        self.show()

    def inicializar_ui(self):
        self.setWindowTitle("Helena Peluqueria")
        self.generar_contenido()

    def generar_contenido(self):
        self.setStyleSheet("background-color: white;")
        self.imagen_fondo_label = QLabel(self)
        pixmap = QPixmap('assets/icono.png')  # Cargar la imagen
        self.imagen_fondo_label.setPixmap(pixmap)
        self.imagen_fondo_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.estructuraExterna()

    def estructuraExterna(self):
        
        self.contenido_central = QGridLayout()
        self.layoutFondo = QGridLayout()
       
        self.layoutFondo.addWidget(QLabel(), 0, 0)
        self.layoutFondo.addWidget(self.imagen_fondo_label, 0, 1)
        self.layoutFondo.addLayout(self.contenido_central, 0, 1)
        self.layoutFondo.addWidget(QLabel(), 0, 2)
        self.layoutFondo.setColumnStretch(0, 1)
        self.layoutFondo.setColumnStretch(1, 8)
        self.layoutFondo.setColumnStretch(2, 1)

        contenedor_widget = QWidget()
        contenedor_widget.setLayout(self.layoutFondo)
        self.setCentralWidget(contenedor_widget)
        self.estructura_central()

    def estructura_central(self):
        self.cargarBarraBusqueda()
        self.cargarTablaClientesyNoData()
        self.cargarEstructuraInferior()
        self.contenido_central.addWidget(self.barraBusqueda, 0, 0)
        self.contenido_central.addWidget(self.tablaClientes, 1, 0)
        self.contenido_central.addWidget(self.no_data_client, 1, 0)
        self.contenido_central.addWidget(self.no_results, 1, 0)
        self.contenido_central.addLayout(self.estructuraInferior,2,0)

        self.contenido_central.setRowStretch(0, 1)
        self.contenido_central.setRowStretch(1, 4)
        self.contenido_central.setRowStretch(2, 5)


    def cargarBarraBusqueda(self):
        self.barraBusqueda = QLineEdit(self)
        self.barraBusqueda.setPlaceholderText("Cercar clients")  # Texto de marcador de posición
        self.barraBusqueda.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.barraBusqueda.textChanged.connect(self.buscar_clientes)

        # Estilo redondeado y con borde naranja
        self.barraBusqueda.setStyleSheet("""
            QLineEdit {
                border: 2px solid orange;
                border-radius: 15px;  /* Cambia este valor para ajustar el redondeado */
                padding: 5px;  /* Añade un poco de espacio interior */
                font-size: 16px;  /* Tamaño de fuente */
            }
        """)

    def cargarTablaClientesyNoData(self):
        self.tablaClientes = QTableWidget(self)
        self.tablaClientes.setColumnCount(6)  # Número de columnas
        self.tablaClientes.setHorizontalHeaderLabels(["Nom", "Cognoms", "Telèfon", "Color"])  # Nombres de las columnas
        self.tablaClientes.setRowCount(0)  # Inicialmente sin filas
        self.configure_table()

        # Conectar el evento de scroll
        self.tablaClientes.verticalScrollBar().valueChanged.connect(self.load_more_clients)
        self.tablaClientes.cellClicked.connect(self.on_cliente_click)

        self.tablaClientes.hideColumn(4)
        self.tablaClientes.hideColumn(5)
        # Contadores para la paginación
        self.offset = 0
        self.limit = 50

        self.no_data_client = QPushButton(self)
        self.no_data_client.setText('No hi han clients afegits')
        self.no_data_client.setStyleSheet("""
            QPushButton {
                font-weight: 100;
                background-color: rgba(211, 211, 211, 0.4);                    
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                font-size: 18px;
                color: black;
                font-weight: bold;
            }
        """)
        self.no_results = QPushButton(self)
        self.no_results.setText("No s'ha trobat cap resultat amb la cerca")
        self.no_results.setStyleSheet("""
            QPushButton {
                font-weight: 100;
                background-color: rgba(211, 211, 211, 0.4);                    
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                font-size: 18px;
                color: black;
                font-weight: bold;
            }
        """)

        svg_renderer = QSvgRenderer('assets/search-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        self.no_results.setIcon(icon)
        self.no_results.setIconSize(svg_pixmap.size())
        self.no_results.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        svg_renderer = QSvgRenderer('assets/information-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        


        self.no_data_client.setIcon(icon)
        self.no_data_client.setIconSize(svg_pixmap.size())
        self.no_data_client.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.no_data_client.hide()
        self.tablaClientes.hide()
        self.no_results.hide()

        # Cargar los primeros datos
        self.buscar_clientes('')

    def configure_table(self):
        header = self.tablaClientes.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)

        # Ocultar el encabezado vertical
        self.tablaClientes.verticalHeader().setVisible(False)

        # Estilo para la tabla
        self.tablaClientes.setStyleSheet("""
            QTableWidget {
                background: transparent;
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                
            }
            QHeaderView::section {
                background-color: rgba(255, 165, 0, 100);
                border: none;
                font-weight: bold;
                font-size: 18px;
                border-radius: 10px;
            }
            QTableWidget::item {
                background-color: rgba(255, 255, 255, 150);
                border: 1px solid #d3d3d3;
            }

            QTableWidget::item:selected {
                background-color: rgba(255, 255, 255, 150); /* Mantener el fondo igual */
                color: black;  /* El color del texto sigue siendo negro (o cualquier color que prefieras) */
            }
            
        """)

    def load_data_clients(self):
        # Conectar a la base de datos
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()

        try:
            # Obtener los datos de los clientes con el ID más alto en orden descendente
            cursor.execute('''
                SELECT Id, Nom, Cognoms, Telefon, Color, Client_antic 
                FROM Client 
                ORDER BY Id DESC 
                LIMIT ? OFFSET ?;
            ''', (self.limit, self.offset))

            rows = cursor.fetchall()

            if len(rows) == 0 and self.offset == 0:
                # Si no hay datos y es la primera carga
                self.tablaClientes.hide()
                self.no_data_client.show()
            else:
                # Si hay datos, añadimos a la tabla
                self.tablaClientes.setRowCount(0)  # Limpia la tabla antes de agregar nuevos datos
                for row in rows:
                    row_position = self.tablaClientes.rowCount()
                    self.tablaClientes.insertRow(row_position)
                    
                    # El ID del cliente (lo vamos a ocultar)
                    id_cliente = row[0]

                    # Iterar por las columnas visibles: Nom, Cognoms, Telefon, Color
                    for column, value in enumerate(row[1:], start=0):  # Empieza desde la primera columna visible
                        item = QTableWidgetItem(str(value))
                        item.setFont(QFont("Arial", 16))

                        # Establecer la celda como no editable
                        item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)

                        # Insertar el valor en la columna correspondiente
                        self.tablaClientes.setItem(row_position, column, item)

                    # Almacenar el ID del cliente en una propiedad oculta
                    self.tablaClientes.setItem(row_position, 4, QTableWidgetItem(str(id_cliente)))

                # Ocultar la columna del ID
                self.tablaClientes.setColumnHidden(4, True)

                # Conectar la acción de clic a la tabla
                self.tablaClientes.cellClicked.connect(self.on_cliente_click)

                # Mostrar la tabla y ocultar el mensaje de "sin datos"
                self.tablaClientes.show()
                self.no_data_client.hide()

        finally:
            # Asegurarse de que la conexión se cierre al final
            conn.close()

    def buscar_clientes(self, texto_buscado):
        # Conectar a la base de datos
        self.offset = 0
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()
        writed=True

        try:
            # Eliminar espacios en blanco innecesarios y descomponer el texto buscado en partes
            texto_buscado = texto_buscado.strip()
            partes = texto_buscado.split()

            if len(partes) == 1:
                # Si hay solo una palabra, buscarla en nombre, apellidos o teléfono
                self.query = '''
                    SELECT Nom, Cognoms, Telefon, Color, Id, Client_antic 
                    FROM Client 
                    WHERE Nom LIKE ? OR Telefon LIKE ?
                    ORDER BY Id DESC
                    LIMIT ? OFFSET ?;
                '''
                params = (f'{texto_buscado}%', f'{texto_buscado}%',f'{self.limit}',f'{self.offset}')
            
            elif len(partes) > 1:
                # La primera palabra es el nombre, el resto son apellidos
                nombre = partes[0]  # Primer palabra es el nombre
                apellidos = ' '.join(partes[1:])  # Todo lo demás son los apellidos

                self.query = '''
                    SELECT Nom, Cognoms, Telefon, Color, Id, Client_antic 
                    FROM Client 
                    WHERE (Nom LIKE ? AND Cognoms LIKE ?) OR (Telefon LIKE ?)
                    ORDER BY Id DESC
                    LIMIT ? OFFSET ?;
                '''
                params = (f'{nombre}%', f'{apellidos}%', f'{texto_buscado}%',f'{self.limit}',f'{self.offset}')

            else:
                # Caso cuando no hay texto de búsqueda (cadena vacía)
                writed=False
                self.query = '''
                    SELECT Nom, Cognoms, Telefon, Color, Id, Client_antic 
                    FROM Client 
                    ORDER BY Id DESC
                    LIMIT ? OFFSET ?;
                '''
                params = (f'{self.limit}',f'{self.offset}')

            self.params_list = list(params)
            cursor.execute(self.query, params)

            rows = cursor.fetchall()

            if len(rows) == 0:
                # Si no hay resultados, ocultar la tabla y mostrar un mensaje
                if not writed:
                    self.tablaClientes.hide()
                    self.no_data_client.show()
                    self.no_results.hide()
                else:

                    self.tablaClientes.hide()
                    self.no_data_client.hide()
                    self.no_results.show()
            else:
                # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                self.tablaClientes.setRowCount(0)
                for row in rows:
                    row_position = self.tablaClientes.rowCount()
                    self.tablaClientes.insertRow(row_position)
                    for column, value in enumerate(row):
                        item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                        item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                        self.tablaClientes.setItem(row_position, column, item)

                self.tablaClientes.show()
                self.no_data_client.hide()
                self.no_results.hide()

        finally:
            # Asegurarse de cerrar la conexión
            conn.close()

    def on_cliente_click(self, row, column):
        # Obtener el ID del cliente de la primera columna (oculta)
        self.id_cliente = self.tablaClientes.item(row, 4).text()
        nom = self.tablaClientes.item(row, 0).text()
        cognoms = self.tablaClientes.item(row, 1).text()
        telefon = self.tablaClientes.item(row, 2).text()
        color = self.tablaClientes.item(row, 3).text()
        tipus_client = self.tablaClientes.item(row, 5).text()
        print(tipus_client)

        if tipus_client == "0":
            self.boto_antic.setIcon(QIcon())
            self.boto_nou.setIcon(self.icon_check)
        elif tipus_client == "1":
            self.boto_antic.setIcon(self.icon_check)
            self.boto_nou.setIcon(QIcon())
        else:
            self.boto_antic.setIcon(QIcon())
            self.boto_nou.setIcon(QIcon())

        self.nom_value_label.setText(nom)
        self.cognoms_value_label.setText(cognoms)
        self.telefon_value_label.setText(telefon)
        self.color_value_label.setText(color)
        self.hide_edit_client()

        self.no_click.hide()
        self.ficha_creacionUsuario.hide()
        self.ficha_usuario.show()
        # Llamar a la función con el ID del cliente
        #self.procesar_cliente(id_cliente)

    def cargarEstructuraInferior(self):
        self.cargarBotonesNuevoyCalculs()
        self.cargarEstructuraDatosUsuario()
        self.cargarCreacionUsuario()
        self.estructuraInferior = QGridLayout()
        self.estructuraInferior.addLayout(self.estructuraInferiorIzquierda,0,0) 
        self.estructuraInferior.addWidget(self.ficha_creacionUsuario,0,1)
        self.estructuraInferior.addWidget(self.ficha_usuario,0,1)
        self.estructuraInferior.addWidget(self.no_click,0,1)
        self.estructuraInferior.setColumnStretch(0,3)
        self.estructuraInferior.setColumnStretch(1,7)



    def cargarBotonesNuevoyCalculs(self):
        create_button = QPushButton(self)
        create_button.setText('Afegir nou client')
        create_button.resize(320 , 120)
        create_button.move(100,400)
        create_button.setStyleSheet(""" QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)
        create_button.clicked.connect(self.crear_usuario)


        svg_renderer = QSvgRenderer('assets/crear-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        self.iconCrear = QIcon(svg_pixmap)

        create_button.setIcon(self.iconCrear)
        create_button.setIconSize(svg_pixmap.size())
        create_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


        
        
        calculs_button = QPushButton(self)
        calculs_button.setText('Càlculs')
        calculs_button.resize(320 , 200)
        calculs_button.move(100,550)
        calculs_button.setStyleSheet(""" QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)
        calculs_button.clicked.connect(self.calculs)


        svg_renderer_calculs = QSvgRenderer('assets/calculator-svg.svg')
        svg_pixmap_calculs = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap_calculs.fill(Qt.GlobalColor.transparent)
        painter_calculs = QPainter(svg_pixmap_calculs)
        svg_renderer_calculs.render(painter_calculs)
        painter_calculs.end()
        icon_calculs = QIcon(svg_pixmap_calculs)

        calculs_button.setIcon(icon_calculs)
        calculs_button.setIconSize(svg_pixmap_calculs.size())
        calculs_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.estructuraInferiorIzquierda = QGridLayout()
        self.estructuraInferiorIzquierda.addWidget(create_button,0,0)
        self.estructuraInferiorIzquierda.addWidget(calculs_button,1,0)
        self.estructuraInferiorIzquierda.setRowStretch(0,5)
        self.estructuraInferiorIzquierda.setRowStretch(1,5)


    def cargarEstructuraDatosUsuario(self):
        self.no_click = QPushButton(self)
        self.no_click.setText("No s'ha seleccionat cap client")
        self.no_click.setStyleSheet("""
            QPushButton {
                font-weight: 100;
                background-color: rgba(211, 211, 211, 0.4);                    
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                font-size: 18px;
                color: black;
                font-weight: bold;
            }
        """)

        svg_renderer = QSvgRenderer('assets/information-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        self.no_click.setIcon(icon)
        self.no_click.setIconSize(svg_pixmap.size())
        self.no_click.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


       

        # Crear un layout dentro del contenedor (como si fuera un div)
        # Widget contenedor del layout
        self.cargarBotonesAcceptaryInput()
        self.ficha_usuario = QWidget(self)  
        # Estilo aplicado al widget que contiene el layout
        self.ficha_usuario.setStyleSheet("""
            background-color: white;
            border: 1px solid #d3d3d3;
        """)
        self.ficha_usuario.hide()

        # Crear el layout principal de la ficha de usuario
        layout_fichaUsuario = QGridLayout(self.ficha_usuario)


        # Añadir widgets dentro del contenedor (QWidget)
        label = QLabel("Fitxa de client")
        label.setStyleSheet("""
            font-weight: 100; /* Peso de la fuente (más fino) */
            background-color: rgba(255, 165, 0, 100);
            border: 1px solid #d3d3d3;
            border-radius: 10px;
            font-weight: bold;  /* Encabezados en negrita */
            font-size: 18px;
            text-align: center;

        """)
        
        #label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding) 
        
        svg_renderer_edit = QSvgRenderer('assets/edit-svg.svg')
        svg_pixmap_edit = QPixmap(24, 24)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap_edit.fill(Qt.GlobalColor.transparent)
        painter_edit = QPainter(svg_pixmap_edit)
        svg_renderer_edit.render(painter_edit)
        painter_edit.end()

        icon_edit = QIcon(svg_pixmap_edit)

        self.edit_name_button = QPushButton(self)
        self.edit_name_button.setText('Editar')
        self.edit_name_button.setStyleSheet(""" QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)
        self.edit_name_button.setIcon(icon_edit)
        self.edit_name_button.setIconSize(svg_pixmap_edit.size())
        self.edit_name_button.clicked.connect(self.editar_nombre)
        self.edit_name_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


        self.edit_surname_button = QPushButton(self)
        self.edit_surname_button.setText('Editar')
        self.edit_surname_button.setStyleSheet(""" QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)
        self.edit_surname_button.setIcon(icon_edit)
        self.edit_surname_button.setIconSize(svg_pixmap_edit.size())
        self.edit_surname_button.clicked.connect(self.editar_apellido)
        self.edit_surname_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


        self.edit_color_button = QPushButton(self)
        self.edit_color_button.setText('Editar')
        self.edit_color_button.setStyleSheet(""" QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)
        self.edit_color_button.setIcon(icon_edit)
        self.edit_color_button.setIconSize(svg_pixmap_edit.size())
        self.edit_color_button.clicked.connect(self.editar_color)
        self.edit_color_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


        self.edit_telefon_button = QPushButton(self)
        self.edit_telefon_button.setText('Editar')
        self.edit_telefon_button.setStyleSheet(""" QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)
        self.edit_telefon_button.setIcon(icon_edit)
        self.edit_telefon_button.setIconSize(svg_pixmap_edit.size())
        self.edit_telefon_button.clicked.connect(self.editar_telefono)
        self.edit_telefon_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)



        informacio_usuari = QGridLayout()

        # Estilo común para los QLabel
        label_style = """
            font-size: 16px;
            font-weight: bold;
            border: none;
        """

        # Añadir los QLabel con el estilo
        nom_label = QLabel("Nom:")
        nom_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(nom_label, 0, 0)

        self.nom_value_label = QLabel()
        self.nom_value_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(self.inputNombre, 0, 1)
        informacio_usuari.addWidget(self.accept_name_button, 0, 2)
        informacio_usuari.addWidget(self.nom_value_label, 0, 1)
        informacio_usuari.addWidget(self.edit_name_button, 0, 2)

        cognoms_label = QLabel("Cognoms:")
        cognoms_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(cognoms_label, 1, 0)

        self.cognoms_value_label = QLabel()
        self.cognoms_value_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(self.inputApellidos, 1, 1)
        informacio_usuari.addWidget(self.accept_surnames_button, 1, 2)
        informacio_usuari.addWidget(self.cognoms_value_label, 1, 1)
        informacio_usuari.addWidget(self.edit_surname_button, 1, 2)

        color_label = QLabel("Color:")
        color_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(color_label, 2, 0)

        self.color_value_label = QLabel()
        self.color_value_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(self.inputColor, 2, 1)
        informacio_usuari.addWidget(self.accept_color_button, 2, 2)
        informacio_usuari.addWidget(self.color_value_label, 2, 1)
        informacio_usuari.addWidget(self.edit_color_button, 2, 2)
        

        telefon_label = QLabel("Telèfon:")
        telefon_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(telefon_label, 3, 0)

        self.telefon_value_label = QLabel()
        self.telefon_value_label.setStyleSheet(label_style)
        informacio_usuari.addWidget(self.inputTelefono, 3, 1)
        informacio_usuari.addWidget(self.accept_phone_button, 3, 2)
        informacio_usuari.addWidget(self.telefon_value_label, 3, 1)
        informacio_usuari.addWidget(self.edit_telefon_button, 3, 2)


        



        informacio_usuari.setColumnStretch(0, 4)  # Columna 0 ocupa el 40%
        informacio_usuari.setColumnStretch(1, 4)  # Columna 1 ocupa el 40%
        informacio_usuari.setColumnStretch(2, 2)

        svg_renderer = QSvgRenderer('assets/tickConfirm-svg.svg')
        svg_pixmap = QPixmap(34, 34)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()

        self.icon_check = QIcon(svg_pixmap)

        layout_eleccion = QGridLayout()
        self.boto_nou = QPushButton("Client nou")
        self.boto_antic = QPushButton("Client antic")

        self.boto_nou.setStyleSheet(""" QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)

        self.boto_antic.setStyleSheet(""" QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """)

        self.boto_antic.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.boto_nou.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.boto_antic.clicked.connect(self.client_antic_click)
        self.boto_nou.clicked.connect(self.client_nou_click)


        layout_eleccion.addWidget(self.boto_antic,0,0)
        layout_eleccion.addWidget(self.boto_nou,0,1)

        layout_tipus_client = QGridLayout()
        label_tipus = QLabel("Tipus de client:")
        label_tipus.setStyleSheet(label_style)
        layout_tipus_client.addWidget(label_tipus, 0, 0)

        layout_tipus_client.addLayout(layout_eleccion, 1, 0)
        layout_tipus_client.setRowStretch(0, 5)  # Primera fila 20%
        layout_tipus_client.setRowStretch(1, 5)
        
        pack_informacio_usuari = QGridLayout()
        pack_informacio_usuari.addLayout(informacio_usuari,0,0)
        pack_informacio_usuari.addLayout(layout_tipus_client,1,0)
        pack_informacio_usuari.setRowStretch(0,7)
        pack_informacio_usuari.setRowStretch(1,3)



        

        svg_renderer_drop = QSvgRenderer('assets/trash-svg.svg')
        svg_pixmap_drop = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap_drop.fill(Qt.GlobalColor.transparent)
        painter_drop = QPainter(svg_pixmap_drop)
        svg_renderer_drop.render(painter_drop)
        painter_drop.end()

        icon_drop = QIcon(svg_pixmap_drop)

        svg_renderer_services = QSvgRenderer('assets/services-svg.svg')
        svg_pixmap_services = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap_services.fill(Qt.GlobalColor.transparent)
        painter_services = QPainter(svg_pixmap_services)
        svg_renderer_services.render(painter_services)
        painter_services.end()

        icon_services = QIcon(svg_pixmap_services)

        serveis_button = QPushButton(self)
        serveis_button.setText('Serveis del client')
        serveis_button.clicked.connect(self.servicios_usuario)
        serveis_button.setIcon(icon_services)
        serveis_button.setIconSize(svg_pixmap_services.size())
        serveis_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding) 
        serveis_button.setStyleSheet("""
            QPushButton {
                font-size: 16px;  /* Tamaño de letra de 16px */
                border: 1px solid black;
                border-radius: 10px;
            }
            QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        """) 



        

        

        grid_layout = QGridLayout()
        grid_layout.addLayout(pack_informacio_usuari, 0, 0)
        grid_layout.addWidget(serveis_button, 0, 1)

        grid_layout.setColumnStretch(0,6)
        grid_layout.setColumnStretch(1,4)
        


        drop_button = QPushButton(self)
        drop_button.setText('Eliminar client')
        drop_button.clicked.connect(self.eliminar_usuario)
        drop_button.setIcon(icon_drop)
        drop_button.setIconSize(svg_pixmap_drop.size())
        drop_button.setStyleSheet("""
            QPushButton {
                font-size: 16px;  /* Tamaño de letra de 16px */
                border: 1px solid black;
                border-radius: 10px;
            }

            QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        """) 
        drop_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)


        

        layout_fichaUsuario.addWidget(label,0,0)
        layout_fichaUsuario.addLayout(grid_layout,1,0)
        layout_fichaUsuario.addWidget(drop_button,2,0)

        
        

        layout_fichaUsuario.setRowStretch(0, 1)  # Primera fila 20%
        layout_fichaUsuario.setRowStretch(1, 7)
        layout_fichaUsuario.setRowStretch(2, 2)

        

    def cargarBotonesAcceptaryInput(self):
        estilosBoton = """ QPushButton{
            
            font-size: 14px;
            border: 1px solid black;
            border-radius: 10px;

        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
    }
        
        """

        estilosInput= """
            QLineEdit {
                border: 2px solid orange;
                border-radius: 15px;  /* Cambia este valor para ajustar el redondeado */
                padding: 5px;  /* Añade un poco de espacio interior */
                font-size: 16px;  /* Tamaño de fuente */
            }
        """
        svg_renderer = QSvgRenderer('assets/tick-svg.svg')
        svg_pixmap = QPixmap(24, 24)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()

        icon = QIcon(svg_pixmap)

        self.accept_name_button = QPushButton(self)
        self.accept_name_button.setText('Acceptar')
        self.accept_name_button.setStyleSheet(estilosBoton)
        self.accept_name_button.setIcon(icon)
        self.accept_name_button.setIconSize(svg_pixmap.size())
        self.accept_name_button.clicked.connect(self.aceptar_nombre)
        self.accept_name_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_name_button.hide()

        self.inputNombre = QLineEdit(self)
        self.inputNombre.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputNombre.setStyleSheet(estilosInput)
        self.inputNombre.returnPressed.connect(self.aceptar_nombre)
        self.inputNombre.hide()


        self.accept_surnames_button = QPushButton(self)
        self.accept_surnames_button.setText('Acceptar')
        self.accept_surnames_button.setStyleSheet(estilosBoton)
        self.accept_surnames_button.setIcon(icon)
        self.accept_surnames_button.setIconSize(svg_pixmap.size())
        self.accept_surnames_button.clicked.connect(self.aceptar_apellido)
        self.accept_surnames_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_surnames_button.hide()

        self.inputApellidos = QLineEdit(self)
        self.inputApellidos.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputApellidos.setStyleSheet(estilosInput)
        self.inputApellidos.returnPressed.connect(self.aceptar_apellido)
        self.inputApellidos.hide()

        self.accept_phone_button = QPushButton(self)
        self.accept_phone_button.setText('Acceptar')
        self.accept_phone_button.setStyleSheet(estilosBoton)
        self.accept_phone_button.setIcon(icon)
        self.accept_phone_button.setIconSize(svg_pixmap.size())
        self.accept_phone_button.clicked.connect(self.aceptar_telefono)
        self.accept_phone_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_phone_button.hide()

        self.inputTelefono = QLineEdit(self)
        self.inputTelefono.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputTelefono.setStyleSheet(estilosInput)
        self.inputTelefono.returnPressed.connect(self.aceptar_telefono)
        self.inputTelefono.hide()

        self.accept_color_button = QPushButton(self)
        self.accept_color_button.setText('Acceptar')
        self.accept_color_button.setStyleSheet(estilosBoton)
        self.accept_color_button.setIcon(icon)
        self.accept_color_button.setIconSize(svg_pixmap.size())
        self.accept_color_button.clicked.connect(self.aceptar_color)
        self.accept_color_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.accept_color_button.hide()

        self.inputColor = QLineEdit(self)
        self.inputColor.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputColor.setStyleSheet(estilosInput)
        self.inputColor.returnPressed.connect(self.aceptar_color)
        self.inputColor.hide()


    def crear_usuario(self):
        self.no_click.hide()
        self.ficha_usuario.hide()
        self.reiniciarInputs()
        #valores inputs a 0
        self.ficha_creacionUsuario.show()

        
    
    def calculs(self):
        ventana_calculos = VentanaCalculs(self)
        ventana_calculos.show()

    def editar_nombre(self):
        self.edit_name_button.hide()
        self.nom_value_label.hide()
        self.inputNombre.setText(self.nom_value_label.text())
        self.inputNombre.show()
        self.accept_name_button.show()
    
    def aceptar_nombre(self):
        nombre = self.inputNombre.text()
        if len(nombre.strip()) == 0:
            ventana_error = VentanaError("El nom no pot estar buit", self)
            ventana_error.exec()
        elif ' ' in nombre:
            ventana_error = VentanaError("Elimina els espais del nom", self)  # Ventana de error si el nombre tiene espacios
            ventana_error.exec()
        else:
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()

            try:
                
                update_query = '''
                    UPDATE Client
                    SET Nom = ?
                    WHERE Id = ?;
                '''
                
                # Ejecutar la query con los parámetros
                cursor.execute(update_query, (nombre, self.id_cliente))
                
                # Confirmar los cambios
                conn.commit()

                self.nom_value_label.setText(nombre)

                self.buscar_clientes(self.barraBusqueda.text())
                
                self.edit_name_button.show()
                self.nom_value_label.show()
                self.inputNombre.hide()
                self.accept_name_button.hide()



                
            except sqlite3.Error as e:
                ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
                ventana_error.exec()
                print(f"Error al actualizar el cliente: {e}")
                
            finally:
                conn.close()


        

    def editar_apellido(self):
        self.edit_surname_button.hide()
        self.cognoms_value_label.hide()
        self.inputApellidos.setText(self.cognoms_value_label.text())
        self.inputApellidos.show()
        self.accept_surnames_button.show()
        

    def aceptar_apellido(self):
        apellidos = self.inputApellidos.text()
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()

        try:
            
            update_query = '''
                UPDATE Client
                SET Cognoms = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(update_query, (apellidos, self.id_cliente))
            
            # Confirmar los cambios
            conn.commit()

            self.cognoms_value_label.setText(apellidos)

            self.buscar_clientes(self.barraBusqueda.text())
            
            self.edit_surname_button.show()
            self.cognoms_value_label.show()
            self.inputApellidos.hide()
            self.accept_surnames_button.hide()



            
        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")
            
        finally:
            conn.close()


    def editar_color(self):
        self.edit_color_button.hide()
        self.color_value_label.hide()
        self.inputColor.setText(self.color_value_label.text())
        self.inputColor.show()
        self.accept_color_button.show()
    
    def aceptar_color(self):
        color = self.inputColor.text()
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()

        try:
            
            update_query = '''
                UPDATE Client
                SET Color = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(update_query, (color, self.id_cliente))
            
            # Confirmar los cambios
            conn.commit()

            self.color_value_label.setText(color)

            self.buscar_clientes(self.barraBusqueda.text())
            
            self.edit_color_button.show()
            self.color_value_label.show()
            self.inputColor.hide()
            self.accept_color_button.hide()



            
        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")
            
        finally:
            conn.close()

    def editar_telefono(self):
        self.edit_telefon_button.hide()
        self.telefon_value_label.hide()
        self.inputTelefono.setText(self.telefon_value_label.text())
        self.inputTelefono.show()
        self.accept_phone_button.show()

    def aceptar_telefono(self):
        telefono = self.inputTelefono.text()
        telefono = telefono.strip()  # Elimina los espacios iniciales y finales
        telefono = telefono.replace(" ", "")
        for caracter in telefono:
            if not caracter.isdigit():
                ventana_error = VentanaError("El telefon nomes pot contindre numeros", self)
                ventana_error.exec()
                return 
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()

        try:
            
            update_query = '''
                UPDATE Client
                SET Telefon = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(update_query, (telefono, self.id_cliente))
            
            # Confirmar los cambios
            conn.commit()

            self.telefon_value_label.setText(telefono)

            self.buscar_clientes(self.barraBusqueda.text())
            
            self.edit_telefon_button.show()
            self.telefon_value_label.show()
            self.inputTelefono.hide()
            self.accept_phone_button.hide()



            
        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haber pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")
            
        finally:
            conn.close()
        



    def hide_edit_client(self):
        self.edit_name_button.show()
        self.nom_value_label.show()
        self.inputNombre.hide()
        self.accept_name_button.hide()
        self.edit_surname_button.show()
        self.cognoms_value_label.show()
        self.inputApellidos.hide()
        self.accept_surnames_button.hide()
        self.edit_color_button.show()
        self.color_value_label.show()
        self.inputColor.hide()
        self.accept_color_button.hide()
        self.edit_telefon_button.show()
        self.telefon_value_label.show()
        self.inputTelefono.hide()
        self.accept_phone_button.hide()

    
    def eliminar_usuario(self):
        dialogo = DialogoConfirmacion(self)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            try:
                conn = sqlite3.connect('data/data.db')  # Conectar a la base de datos
                cursor = conn.cursor()

                # Supongamos que `self.id_cliente` contiene el ID del cliente a eliminar
                delete_query = 'DELETE FROM Client WHERE Id = ?'
                cursor.execute(delete_query, (self.id_cliente,))

                # Confirmar cambios en la base de datos
                conn.commit()

                self.ficha_usuario.hide() 
                self.no_click.show()

                self.buscar_clientes(self.barraBusqueda.text())


            except sqlite3.Error as e:
                print(f"Error al eliminar el cliente: {e}")

            finally:
                conn.close()

    def servicios_usuario(self):
        ventana_servicios = VentanaServicios(self)
        ventana_servicios.show()
        
        
    
    def hide_layout(self,layout):
        for i in range(layout.count()):
            widget = layout.itemAt(i).widget()
            if widget is not None:
                widget.hide()
    def show_layout(self,layout):
        for i in range(layout.count()):
            widget = layout.itemAt(i).widget()
            if widget is not None:
                widget.show()
    def client_antic_click(self):
        
        if self.boto_antic.icon().cacheKey()!= 0:
            self.actualizar_cliente_tipo(None)
            self.boto_antic.setIcon(QIcon())

        else:
            self.actualizar_cliente_tipo(1)
            self.boto_nou.setIcon(QIcon())
            self.boto_antic.setIcon(self.icon_check)
            
            # Funcionalidad cuando está desactivado
    def client_nou_click(self):
        if self.boto_nou.icon().cacheKey() != 0:
            self.actualizar_cliente_tipo(None)
            self.boto_nou.setIcon(QIcon())

        else:
            self.actualizar_cliente_tipo(0)
            self.boto_antic.setIcon(QIcon())
            self.boto_nou.setIcon(self.icon_check)
        # Funcionalidad cuando está desactivado

    
    def actualizar_cliente_tipo(self,client_antic):
        conn = sqlite3.connect('data/data.db')  # Conectar a la base de datos
        cursor = conn.cursor()

        try:
            # Consulta para actualizar tanto Telefon como Client_antic
            update_query = '''
                UPDATE Client
                SET Client_antic = ?
                WHERE Id = ?;
            '''
            
            # Ejecutar la query con los parámetros (telefono y client_antic)
            cursor.execute(update_query, (client_antic, self.id_cliente))
            
            # Confirmar los cambios
            conn.commit()

            # Actualizar la interfaz de usuario
            self.buscar_clientes(self.barraBusqueda.text())


        except sqlite3.Error as e:
            ventana_error = VentanaError("Alguna cosa no tenia que haver pasat, parla amb Martí sobre aquest error", self)
            ventana_error.exec()
            print(f"Error al actualizar el cliente: {e}")

        finally:
            # Cerrar la conexión a la base de datos
            conn.close()

    

    

    def load_more_clients(self):
        if self.tablaClientes.verticalScrollBar().value() == self.tablaClientes.verticalScrollBar().maximum():

            self.offset += self.limit
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            self.params_list[-1] = f'{self.offset}'
            params = tuple(self.params_list)
            try:
                cursor.execute(self.query, params)

                rows = cursor.fetchall()

                if len(rows) == 0:
                    pass
                else:
                    # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                    for row in rows:
                        row_position = self.tablaClientes.rowCount()
                        self.tablaClientes.insertRow(row_position)
                        for column, value in enumerate(row):
                            item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                            item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                            self.tablaClientes.setItem(row_position, column, item)

                
            finally:
                # Asegurarse de cerrar la conexión
                conn.close()

    def cargarCreacionUsuario(self):
        self.ficha_creacionUsuario = QWidget(self)  
        # Estilo aplicado al widget que contiene el layout
        self.ficha_creacionUsuario.setStyleSheet("""
            background-color: white;
            border: 1px solid #d3d3d3;
        """)
        self.ficha_creacionUsuario.hide()
        
        estilosInput= """
            QLineEdit {
                border: 2px solid orange;
                border-radius: 15px;  /* Cambia este valor para ajustar el redondeado */
                padding: 5px;  /* Añade un poco de espacio interior */
                font-size: 16px;  /* Tamaño de fuente */
            }
        """

        label_style = """
            font-size: 16px;
            font-weight: bold;
            border: none;
        """

        estiloBoton= """ QPushButton{
            
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;


        } 
        
        QPushButton:hover {
        background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
        border: 2px solid #FFA500;  /* Borde naranja */
        color: white;  /* Cambiar el color del texto a blanco */
          /* Cambiar el cursor a una mano */
        }
        
        """

        # Crear el layout principal de la ficha de usuario
        layout_fichaCreacionsuario = QGridLayout(self.ficha_creacionUsuario)


        # Añadir widgets dentro del contenedor (QWidget)
        label = QLabel("Creació de client")
        label.setStyleSheet("""
            font-weight: 100; /* Peso de la fuente (más fino) */
            background-color: rgba(255, 165, 0, 100);
            border: 1px solid #d3d3d3;
            border-radius: 10px;
            font-weight: bold;  /* Encabezados en negrita */
            font-size: 18px;
            text-align: center;

        """)
        layout_datosUsuario = QGridLayout()
        layout_abajo = QGridLayout()

        nom_label = QLabel("Nom:")
        nom_label.setStyleSheet(label_style)
        nom_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        cognoms_label = QLabel("Cognoms:")
        cognoms_label.setStyleSheet(label_style)
        cognoms_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        color_label = QLabel("Color:")
        color_label.setStyleSheet(label_style)
        color_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        telefon_label = QLabel("Telèfon:")
        telefon_label.setStyleSheet(label_style)
        telefon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        

        self.inputNombreCreacion = QLineEdit(self)
        self.inputNombreCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputNombreCreacion.setStyleSheet(estilosInput)

        self.inputApellidosCreacion = QLineEdit(self)
        self.inputApellidosCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputApellidosCreacion.setStyleSheet(estilosInput)

        self.inputColorCreacion = QLineEdit(self)
        self.inputColorCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputColorCreacion.setStyleSheet(estilosInput)

        self.inputTelefonoCreacion = QLineEdit(self)
        self.inputTelefonoCreacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.inputTelefonoCreacion.setStyleSheet(estilosInput)

        layout_datosUsuario.addWidget(nom_label,0,0)
        layout_datosUsuario.addWidget(self.inputNombreCreacion,0,1)
        layout_datosUsuario.addWidget(cognoms_label,1,0)
        layout_datosUsuario.addWidget(self.inputApellidosCreacion,1,1)
        layout_datosUsuario.addWidget(color_label,2,0)
        layout_datosUsuario.addWidget(self.inputColorCreacion,2,1)
        layout_datosUsuario.addWidget(telefon_label,3,0)
        layout_datosUsuario.addWidget(self.inputTelefonoCreacion,3,1)
        layout_datosUsuario.setColumnStretch(0,5)
        layout_datosUsuario.setColumnStretch(1,5)

        layout_eleccion = QGridLayout()
        self.boto_Creacionou = QPushButton("Client nou")
        self.boto_Creacioantic = QPushButton("Client antic")

        self.boto_Creacionou.setStyleSheet(estiloBoton)

        self.boto_Creacioantic.setStyleSheet(estiloBoton)

        self.boto_Creacioantic.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.boto_Creacionou.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.boto_Creacioantic.clicked.connect(self.clientCreacio_antic_click)
        self.boto_Creacionou.clicked.connect(self.clientCreacio_nou_click)


        layout_eleccion.addWidget(self.boto_Creacioantic,0,0)
        layout_eleccion.addWidget(self.boto_Creacionou,0,1)

        layout_tipus_client = QGridLayout()
        label_tipus = QLabel("Tipus de client:")
        label_tipus.setStyleSheet(label_style)
        layout_tipus_client.addWidget(label_tipus, 0, 0)

        layout_tipus_client.addLayout(layout_eleccion, 1, 0)
        layout_tipus_client.setRowStretch(0, 5)  # Primera fila 20%
        layout_tipus_client.setRowStretch(1, 5)

        pack_informacio_usuari = QGridLayout()
        pack_informacio_usuari.addLayout(layout_datosUsuario,0,0)
        pack_informacio_usuari.addLayout(layout_tipus_client,1,0)
        pack_informacio_usuari.setRowStretch(0,7)
        pack_informacio_usuari.setRowStretch(1,3)

        aceptar_btn = QPushButton('Crear')
        svg_pixmap = QPixmap(44, 44)  # Elige el tamaño que quieras para tu ícono
        aceptar_btn.setIcon(self.iconCrear)
        aceptar_btn.setIconSize(svg_pixmap.size())
        aceptar_btn.setStyleSheet(estiloBoton)
        aceptar_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        aceptar_btn.clicked.connect(self.aceptarCreacion)

        cancelar_btn = QPushButton('Cancelar')
        svg_renderer = QSvgRenderer('assets/cross-svg.svg')
        svg_pixmap = QPixmap(44, 44)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        cancelar_btn.setIcon(icon)
        cancelar_btn.setIconSize(svg_pixmap.size())
        cancelar_btn.setStyleSheet(estiloBoton)
        cancelar_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        cancelar_btn.clicked.connect(self.rechazar_creacion)

        layout_abajo.addWidget(aceptar_btn,0,0)
        layout_abajo.addWidget(cancelar_btn,0,1)
        layout_abajo.setColumnStretch(0,5)
        layout_abajo.setColumnStretch(1,5)




        layout_fichaCreacionsuario.addWidget(label,0,0)
        layout_fichaCreacionsuario.addLayout(pack_informacio_usuari,1,0)
        label_sin_bordes = QLabel("")
        label_sin_bordes.setStyleSheet("border: none;") 
        layout_fichaCreacionsuario.addWidget(label_sin_bordes, 2, 0)
        layout_fichaCreacionsuario.addLayout(layout_abajo,3,0)


        layout_fichaCreacionsuario.setRowStretch(0,2)
        layout_fichaCreacionsuario.setRowStretch(1,14)
        layout_fichaCreacionsuario.setRowStretch(2,1)
        layout_fichaCreacionsuario.setRowStretch(3,3)
    
    def aceptarCreacion(self):
        nombre = self.inputNombreCreacion.text()
        cognoms = self.inputApellidosCreacion.text()
        color = self.inputColorCreacion.text()
        telefon = self.inputTelefonoCreacion.text()

        if len(nombre.strip()) == 0:
            ventana_error = VentanaError("El nom no pot estar buit", self)
            ventana_error.exec()
            return
        if ' ' in nombre:
            ventana_error = VentanaError("Elimina els espais del nom", self)  # Ventana de error si el nombre tiene espacios
            ventana_error.exec()
            return
        telefono = telefon.strip()  # Elimina los espacios iniciales y finales
        telefono = telefono.replace(" ", "")
        for caracter in telefono:
            if not caracter.isdigit():
                ventana_error = VentanaError("El telefon nomes pot contindre numeros", self)
                ventana_error.exec()
                return 

        if self.boto_Creacioantic.icon().cacheKey() !=0:
            client_antic=1
        elif self.boto_Creacionou.icon().cacheKey() !=0:
            client_antic=0
        else:
            client_antic= None

        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            
            # Query de inserción
            insert_query = '''
                INSERT INTO Client (Nom, Cognoms, Telefon, Client_antic, Color)
                VALUES (?, ?, ?, ?, ?);
            '''
            
            # Ejecutar la query con los parámetros
            cursor.execute(insert_query, (nombre, cognoms, telefono, client_antic, color))
            
            # Confirmar los cambios
            conn.commit()
            
            self.barraBusqueda.setText("")
            self.buscar_clientes("")

            self.on_cliente_click(0,0)



            print("El cliente ha sido insertado con éxito.")
        
        except sqlite3.Error as e:
            print(f"Error al insertar el cliente: {e}")
            
        finally:
            # Cerrar la conexión
            conn.close()
            
            print("introduciendo datos")


    def clientCreacio_antic_click(self):
        if self.boto_Creacioantic.icon().cacheKey()!= 0:
            self.boto_Creacioantic.setIcon(QIcon())

        else:
            self.boto_Creacionou.setIcon(QIcon())
            self.boto_Creacioantic.setIcon(self.icon_check)

        

    def clientCreacio_nou_click(self):
        if self.boto_Creacionou.icon().cacheKey() != 0:
            self.boto_Creacionou.setIcon(QIcon())

        else:
            self.boto_Creacioantic.setIcon(QIcon())
            self.boto_Creacionou.setIcon(self.icon_check)


    def reiniciarInputs(self):
        self.inputNombreCreacion.setText("")
        self.inputApellidosCreacion.setText("")
        self.inputColorCreacion.setText("")
        self.inputTelefonoCreacion.setText("")
        self.boto_Creacioantic.setIcon(QIcon())
        self.boto_Creacionou.setIcon(QIcon())

    def rechazar_creacion(self):
        self.ficha_creacionUsuario.hide()
        self.ficha_usuario.hide()  
        self.no_click.show()

class VentanaCalculs(QMainWindow):
    def __init__(self,parent= None):
        super().__init__(parent)


        # Configurar la ventana
        self.setWindowTitle("Càlculs")

        self.generar_contenido()
        self.show()

    def generar_contenido(self):
        estiloBoton = """ QPushButton{
            font-size: 16px;
            border: 1px solid black;
            border-radius: 10px;
        } 
            
        QPushButton:hover {
            background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
            border: 2px solid #FFA500;  /* Borde naranja */
            color: white;  /* Cambiar el color del texto a blanco */
            /* Cambiar el cursor a una mano */
        }"""

        layoutFondo = QGridLayout()

        # Crear un frame para el botón Tipo Cliente
        frameTipoCliente = QFrame()
        frameTipoCliente.setStyleSheet("padding: 10px; margin: 5px;")  # Padding y margen
        botonTipoCliente = QPushButton('Tipus de clients')
        svg_renderer = QSvgRenderer('assets/typeuser-svg.svg')
        svg_pixmap = QPixmap(44, 44)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        botonTipoCliente.setIcon(icon)
        botonTipoCliente.setIconSize(svg_pixmap.size())
        botonTipoCliente.setStyleSheet(estiloBoton)
        botonTipoCliente.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        botonTipoCliente.clicked.connect(self.tipus_client_function)
        frameTipoCliente.setLayout(QVBoxLayout())
        frameTipoCliente.layout().addWidget(botonTipoCliente)
        layoutFondo.addWidget(frameTipoCliente, 0, 0)

        # Espaciador
        layoutFondo.addWidget(QLabel(), 1, 0)

        # Crear un frame para el botón Facturación
        frameFacturacion = QFrame()
        frameFacturacion.setStyleSheet("padding: 10px; margin: 5px;")  # Padding y margen
        botonFacturacion = QPushButton('Facturació total')
        svg_renderer = QSvgRenderer('assets/euro-svg.svg')
        svg_pixmap = QPixmap(44, 44)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        botonFacturacion.setIcon(icon)
        botonFacturacion.setIconSize(svg_pixmap.size())
        botonFacturacion.setStyleSheet(estiloBoton)
        botonFacturacion.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        botonFacturacion.clicked.connect(self.generar_facturacio)
        frameFacturacion.setLayout(QVBoxLayout())
        frameFacturacion.layout().addWidget(botonFacturacion)
        layoutFondo.addWidget(frameFacturacion, 2, 0)

        # Espaciador
        layoutFondo.addWidget(QLabel(), 3, 0)

        # Crear un frame para el botón Fidelidad
        frameFidelidad = QFrame()
        frameFidelidad.setStyleSheet("padding: 10px; margin: 5px;")  # Padding y margen
        botonFidelidad = QPushButton('Taula de fidelitat')
        svg_renderer = QSvgRenderer('assets/heart-svg.svg')
        svg_pixmap = QPixmap(44, 44)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)

        botonFidelidad.setIcon(icon)
        botonFidelidad.setIconSize(svg_pixmap.size())
        botonFidelidad.setStyleSheet(estiloBoton)
        botonFidelidad.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        botonFidelidad.clicked.connect(self.generar_fidelidad)
        frameFidelidad.setLayout(QVBoxLayout())
        frameFidelidad.layout().addWidget(botonFidelidad)
        layoutFondo.addWidget(frameFidelidad, 4, 0)

        layoutFondo.setRowStretch(0, 2)
        layoutFondo.setRowStretch(1, 1)
        layoutFondo.setRowStretch(2, 2)
        layoutFondo.setRowStretch(3, 1)
        layoutFondo.setRowStretch(4, 2)

        layoutCentral = QGridLayout()
        layoutCentral.addWidget(QLabel(), 0, 0)
        layoutCentral.addLayout(layoutFondo, 0, 1)
        layoutCentral.addWidget(QLabel(), 0, 2)
        layoutCentral.setColumnStretch(0, 2)
        layoutCentral.setColumnStretch(1, 6)
        layoutCentral.setColumnStretch(2, 2)

        contenedor_widget = QWidget()
        contenedor_widget.setLayout(layoutCentral)
        self.setCentralWidget(contenedor_widget)


    def tipus_client_function(self):
        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()

            # Seleccionar el campo 'tipus_client' de todos los clientes
            cursor.execute('SELECT Client_antic FROM Client')
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Guardar los valores de 'tipus_client' en una lista
            tipus_client_list = [fila[0] for fila in resultados]

            # Cerrar la conexión a la base de datos
            conn.close()
            cantidad_total = len(tipus_client_list)
        


            if cantidad_total > 1:
                cantidad_antiguos = tipus_client_list.count(1)
                cantidad_nuevos = tipus_client_list.count(0)
                cantidad_noDefinidos = tipus_client_list.count(None)
                perc_antiguos = round((cantidad_antiguos / cantidad_total) * 100)
                perc_nuevos = round((cantidad_nuevos / cantidad_total) * 100)
                perc_noDefinidos = round((cantidad_noDefinidos / cantidad_total) * 100)
                
                ventana_tipoCliente = VentanaTipusClients(str(perc_antiguos),str(perc_noDefinidos),str(perc_nuevos),str(cantidad_antiguos),str(cantidad_nuevos),str(cantidad_noDefinidos),str(cantidad_total),self)
                ventana_tipoCliente.show()
                


            else:
                ventana_error = VentanaError("No hi han suficients clients", self)
                ventana_error.exec()

            # Devolver la lista de tipus_client

        except sqlite3.Error as e:
            # Manejar cualquier error que ocurra durante la operación con la base de datos
            print(f"Ha ocurrido un error con la base de datos: {e}")

        except Exception as e:
            # Capturar cualquier otro tipo de excepción inesperada
            print(f"Ha ocurrido un error inesperado: {e}")
        
    def generar_facturacio(self):
        ventana_facturacio = VentanaFacturacio(self)
        ventana_facturacio.show()
    def generar_fidelidad(self):
        ventana_fidelidad = VentanaFidelitat(self)
        ventana_fidelidad.show()

        


    def click_generar(self):
        try:
            # Convertir ambas fechas a objetos datetime para poder compararlas
            fecha1 = datetime.strptime(self.input1_content, "%d/%m/%Y")
            fecha2 = datetime.strptime(self.input2_content, "%d/%m/%Y")
            

            # Comparar si la primera fecha es menor que la segunda
            if fecha1 < fecha2:
                self.cargar_datos()
            else:
                ventana_error = VentanaError("La primera data ha de ser anterior a la segona", self)
                ventana_error.exec()

        except ValueError:
            # Capturar el error si los inputs no son fechas válidas
            ventana_error = VentanaError("La data no té un format vàlid", self)
            ventana_error.exec()        

    def cargar_datos(self):
        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            #print(fecha1)
            #print(fecha2)

            # Seleccionar los campos 'Servei', 'Fecha' y 'Preu' de la tabla 'Serveis'
            cursor.execute('''
            SELECT Servei, Fecha, Preu 
            FROM Serveis 
            WHERE Fecha BETWEEN ? AND ?
        ''', (self.input1_content, self.input2_content))
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            if not resultados:
                ventana_error = VentanaError("No hi han registres de serveis en aquesta data", self)
                ventana_error.exec()
                return 
            
            # Guardar los datos en una lista o hacer algo con ellos
            servicios = []
            for row in resultados:
                # Cada fila contiene ('Servei', 'Fecha', 'Preu')
                servicio = {
                    'Servei': row[0],
                    'Fecha': row[1],
                    'Preu': row[2]
                }
                servicios.append(servicio)

            # Devolver o procesar los datos
            self.construirExcel(servicios)

        except sqlite3.Error as e:
            print(f"Error al acceder a la base de datos: {e}")

        finally:
            # Cerrar la conexión a la base de datos
            if conn:
                conn.close()

    def construirExcel(self, servicios, nombre_archivo='servicios.xlsx'):
        # Crear un nuevo libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active

        # Añadir cabeceras para la primera parte del Excel

        font_bold = Font(bold=True)
        fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        headers = ["Data", "Servei", "Preu", "Total"]
        ws.append(headers)

        # Aplicar el formato a las cabeceras (fondo naranja y negrita)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_bold
            cell.fill = fill_orange

        total_precio = 0  # Inicializar el total de precios

        # Añadir los datos de los servicios
        for servicio in servicios:
            fecha = servicio['Fecha']
            servei = servicio['Servei']
            preu = servicio['Preu']
            total_precio += preu
            preu_simbolo = f"{preu} €"
            # Añadir una nueva fila con la fecha, el servicio y el precio
            ws.append([fecha, servei, preu_simbolo, ""])

        # Colocar el total en la cuarta columna, en negrita
        total_cell = ws.cell(row=ws.max_row + 1, column=4, value=f"{total_precio} €")
        total_label_cell = ws.cell(row=ws.max_row, column=3, value="Total")
        total_label_cell.font = Font(bold=True)
        total_cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 15  # Ajustar el ancho de la columna A (Fecha)
        ws.column_dimensions['B'].width = 30  # Ajustar el ancho de la columna B (Servei)
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        

        # Obtener los datos de "Serveis oferits" y contar los servicios coincidentes
        

        # Guardar el archivo Excel con el nombre especificado
        wb.save(nombre_archivo)
        print(f"Archivo Excel '{nombre_archivo}' creado con éxito.")

        

class VentanaTipusClients(QMainWindow):
    def __init__(self, perc_antiguos,perc_noDefinidos, perc_nuevos, cantidad_antiguos, cantidad_nuevos,cantidad_noDefinidos, cantidad_total, parent=None):
        super().__init__(parent)


        # Configurar la ventana
        self.setWindowTitle("Tipus de clients")

        self.generar_contenido(perc_antiguos,perc_noDefinidos,perc_nuevos,cantidad_antiguos,cantidad_nuevos,cantidad_noDefinidos,cantidad_total)
        self.parent_window = parent
        self.parent_window.close()
        self.show()
    def generar_contenido(self,perc_antiguos,perc_noDefinidos,perc_nuevos,cantidad_antiguos,cantidad_nuevos,cantidad_noDefinidos,cantidad_total):
        # Estilo para los labels
        label_style = """
            font-size: 16px;
            font-weight: bold;
            border: none;
        """

        # Estilo para los frames
        frame_style = """
            border: 2px solid rgba(255, 165, 0, 100);  /* Borde naranja */
            border-radius: 5px;         /* Bordes redondeados opcional */
            padding: 10px;               /* Espaciado interno opcional */
        """

        layout_superior = QGridLayout()
        label_clients_antics = QLabel(perc_antiguos+'% Clients Antics')
        label_clients_antics.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_superior.addWidget(label_clients_antics, 0, 0)

        label_clients_nous = QLabel(perc_nuevos+'% Clients Nous')
        label_clients_nous.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_superior.addWidget(label_clients_nous, 1, 0)

        label_clients_noDefinidos = QLabel(perc_noDefinidos+'% Clients No Definits')
        label_clients_noDefinidos.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_superior.addWidget(label_clients_noDefinidos, 2, 0)

        # Crear un frame para el layout superior
        frame_superior = QFrame()
        frame_superior.setLayout(layout_superior)
        frame_superior.setStyleSheet(frame_style)  # Aplicar estilo al frame

        # Establecer márgenes y espaciado para el layout superior
        layout_superior.setContentsMargins(10, 10, 10, 10)  # (izquierda, arriba, derecha, abajo)
        layout_superior.setSpacing(5)  # Espaciado entre widgets

        layout_central = QGridLayout()
        label_clients_antics_count = QLabel(cantidad_antiguos+' Clients Antics')
        label_clients_antics_count.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_central.addWidget(label_clients_antics_count, 0, 0)

        label_clients_nous_count = QLabel(cantidad_nuevos+' Clients Nous')
        label_clients_nous_count.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_central.addWidget(label_clients_nous_count, 1, 0)

        label_clients_noDefinits = QLabel(cantidad_noDefinidos+' Clients No Definits')
        label_clients_noDefinits.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_central.addWidget(label_clients_noDefinits, 2, 0)

        # Crear un frame para el layout central
        frame_central = QFrame()
        frame_central.setLayout(layout_central)
        frame_central.setStyleSheet(frame_style)  # Aplicar estilo al frame

        # Establecer márgenes y espaciado para el layout central
        layout_central.setContentsMargins(10, 10, 10, 10)  # (izquierda, arriba, derecha, abajo)
        layout_central.setSpacing(5)  # Espaciado entre widgets

        layout_inferior = QGridLayout()
        label_clients_totals = QLabel(cantidad_total+' Clients Totals')
        label_clients_totals.setStyleSheet(label_style)  # Aplicar estilo al QLabel
        layout_inferior.addWidget(label_clients_totals, 0, 0)

        # Crear un frame para el layout inferior
        frame_inferior = QFrame()
        frame_inferior.setLayout(layout_inferior)
        frame_inferior.setStyleSheet(frame_style)  # Aplicar estilo al frame

        # Establecer márgenes y espaciado para el layout inferior
        layout_inferior.setContentsMargins(10, 10, 10, 10)  # (izquierda, arriba, derecha, abajo)
        layout_inferior.setSpacing(5)  # Espaciado entre widgets

        layoutCentral = QGridLayout()
        layoutCentral.addWidget(frame_superior, 0, 0)
        layoutCentral.addWidget(frame_central, 1, 0)
        layoutCentral.addWidget(frame_inferior, 2, 0)

        # Establecer márgenes y espaciado para el layout principal
        layoutCentral.setContentsMargins(20, 20, 20, 20)  # (izquierda, arriba, derecha, abajo)
        layoutCentral.setSpacing(10)  # Espaciado entre frames

        contenedor_widget = QWidget()
        contenedor_widget.setLayout(layoutCentral)
        self.setCentralWidget(contenedor_widget)

class VentanaFacturacio(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Facturació")

        self.generar_contenido()
        self.parent_window = parent
        self.parent_window.close()
        self.show()

    def generar_contenido(self):
        label_style = """
            font-size: 16px;
            font-weight: bold;
            border: none;
        """

        # Estilo para los frames
        frame_style = """
            QFrame {
                border: 2px solid orange;
                border-radius: 5px;
                padding: 10px;
                background-color: #fff8e1;  /* Fondo claro opcional */
            }
        """
        
        # Estilos para los inputs
        estilosInput = """
            QLineEdit {
                border: 2px solid orange;
                border-radius: 15px;
                padding: 5px;
                font-size: 16px;
            }
        """
        
        # Estilos para los botones
        estilosBoton = """
            QPushButton {
                font-size: 16px;
                border: 1px solid black;
                border-radius: 10px;
                padding: 10px 20px;
            }
            
            QPushButton:hover {
                background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
                border: 2px solid #FFA500;  /* Borde naranja */
                color: white;  /* Cambiar el color del texto a blanco */
            }
        """

        # Crear labels
        label_desde = QLabel("Desde")
        label_desde.setStyleSheet(label_style)

        label_fins = QLabel("Fins a")
        label_fins.setStyleSheet(label_style)

        # Crear inputs
        self.input1 = QLineEdit(self)
        self.input1.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.input1.setStyleSheet(estilosInput)

        self.input2 = QLineEdit(self)
        self.input2.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.input2.setStyleSheet(estilosInput)

        # Crear botones
        generar_btn = QPushButton('Generar facturació')
        generar_btn.setStyleSheet(estilosBoton)
        generar_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        generar_btn.clicked.connect(self.click_generar)


        self.descargar_btn = QPushButton('Baixar excel de facturació')
        self.descargar_btn.setStyleSheet(estilosBoton)
        self.descargar_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Añadir icono al botón de descarga
        svg_renderer = QSvgRenderer('assets/download-svg.svg')
        svg_pixmap = QPixmap(44, 44)  # Tamaño del ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        icon = QIcon(svg_pixmap)
        self.descargar_btn.setIcon(icon)
        self.descargar_btn.setIconSize(svg_pixmap.size())
        self.descargar_btn.clicked.connect(self.click_instalar)
        self.descargar_btn.hide()


        # Crear frames
        frame_izquierda = QFrame()
        frame_izquierda.setStyleSheet(frame_style)
        layout_izquierda = QGridLayout()
        layout_izquierda.addWidget(label_desde, 0, 0)
        layout_izquierda.addWidget(self.input1, 0, 1)
        layout_izquierda.setContentsMargins(10, 10, 10, 10)  # Añadir márgenes internos
        layout_izquierda.setSpacing(15)  # Añadir espacio entre los widgets
        frame_izquierda.setLayout(layout_izquierda)

        frame_derecha = QFrame()
        frame_derecha.setStyleSheet(frame_style)
        layout_derecha = QGridLayout()
        layout_derecha.addWidget(label_fins, 0, 0)
        layout_derecha.addWidget(self.input2, 0, 1)
        layout_derecha.setContentsMargins(10, 10, 10, 10)  # Añadir márgenes internos
        layout_derecha.setSpacing(15)  # Añadir espacio entre los widgets
        frame_derecha.setLayout(layout_derecha)

        # Layout superior para las dos secciones
        layout_superior = QGridLayout()
        layout_superior.addWidget(frame_izquierda, 0, 0)
        layout_superior.addWidget(frame_derecha, 0, 1)
        layout_superior.setContentsMargins(20, 20, 20, 20)  # Márgenes alrededor del layout superior
        layout_superior.setSpacing(20)  # Espaciado entre los frames

        # Layout principal
        layoutCentral = QGridLayout()
        layoutCentral.addLayout(layout_superior, 0, 0)
        layoutCentral.addWidget(generar_btn, 1, 0)
        layoutCentral.addWidget(self.descargar_btn, 2, 0)
        layoutCentral.setContentsMargins(30, 30, 30, 30)  # Márgenes alrededor de todo el layout central
        layoutCentral.setSpacing(20)  # Espacio entre botones y otros elementos

        # Establecer el layout central en un widget contenedor
        self.generar_fechas()
        contenedor_widget = QWidget()
        contenedor_widget.setLayout(layoutCentral)
        self.setCentralWidget(contenedor_widget)

    def string_a_fecha(self,fecha_str):
        return datetime.strptime(fecha_str, '%d/%m/%Y').date()

    # Conversión de objeto date a string
    def fecha_a_string(self,fecha_str):
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')  # Convierte la cadena a objeto datetime
        return fecha_obj.strftime('%d/%m/%Y')

    def generar_fechas(self):
        dia_actual = datetime.now()

        # Formatear la fecha actual en formato dd/mm/yyyy
        self.input2.setText(dia_actual.strftime("%d/%m/%Y"))

        # Obtener el primer día del mes actual
        primer_dia_mes = dia_actual.replace(day=1)
        
        # Formatear el primer día del mes en formato dd/mm/yyyy
        self.input1.setText(primer_dia_mes.strftime("%d/%m/%Y"))

        

        #Eliminar ventana padre


    def click_generar(self):
        try:
            # Convertir ambas fechas a objetos datetime para poder compararlas
            fecha1 = datetime.strptime(self.input1.text(), "%d/%m/%Y")
            fecha2 = datetime.strptime(self.input2.text(), "%d/%m/%Y")
            

            # Comparar si la primera fecha es menor que la segunda
            if fecha1 < fecha2:
                self.input1_content = self.string_a_fecha(self.input1.text())
                self.input2_content = self.string_a_fecha(self.input2.text())

                self.cargar_datos()
            else:
                ventana_error = VentanaError("La primera data ha de ser anterior a la segona", self)
                ventana_error.exec()

        except ValueError:
            # Capturar el error si los inputs no son fechas válidas
            ventana_error = VentanaError("La data no té un format vàlid", self)
            ventana_error.exec()        

    def cargar_datos(self):
        try:
            # Conectar a la base de datos
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            #print(fecha1)
            #print(fecha2)

            # Seleccionar los campos 'Servei', 'Fecha' y 'Preu' de la tabla 'Serveis'
            cursor.execute('''
            SELECT Servei, Fecha, Preu 
            FROM Serveis 
            WHERE Fecha BETWEEN ? AND ?
        ''', (self.input1_content, self.input2_content))
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            if not resultados:
                ventana_error = VentanaError("No hi han registres de serveis en aquesta data", self)
                ventana_error.exec()
                return 
            
            # Guardar los datos en una lista o hacer algo con ellos
            servicios = []
            for row in resultados:
                # Cada fila contiene ('Servei', 'Fecha', 'Preu')
                servicio = {
                    'Servei': row[0],
                    'Fecha': row[1],
                    'Preu': row[2]
                }
                servicios.append(servicio)

            # Devolver o procesar los datos
            self.construirExcel(servicios)

        except sqlite3.Error as e:
            print(f"Error al acceder a la base de datos: {e}")

        finally:
            # Cerrar la conexión a la base de datos
            if conn:
                conn.close()

    def construirExcel(self, servicios, nombre_archivo='excels/facturacio.xlsx'):
        # Crear un nuevo libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active

        # Añadir cabeceras para la primera parte del Excel

        font_bold = Font(bold=True)
        fill_orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        headers = ["Data", "Servei", "Preu", "Total"]
        ws.append(headers)

        # Aplicar el formato a las cabeceras (fondo naranja y negrita)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_bold
            cell.fill = fill_orange

        total_precio = 0  # Inicializar el total de precios

        # Añadir los datos de los servicios
        for servicio in servicios:
            fecha = self.fecha_a_string(servicio['Fecha'])
            servei = servicio['Servei']
            preu = servicio['Preu']
            total_precio += preu
            preu_simbolo = f"{preu} €"
            # Añadir una nueva fila con la fecha, el servicio y el precio
            ws.append([fecha, servei, preu_simbolo, ""])

        # Colocar el total en la cuarta columna, en negrita
        total_cell = ws.cell(row=ws.max_row + 1, column=4, value=f"{total_precio} €")
        total_label_cell = ws.cell(row=ws.max_row, column=3, value="Total")
        total_label_cell.font = Font(bold=True)
        total_cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 15  # Ajustar el ancho de la columna A (Fecha)
        ws.column_dimensions['B'].width = 30  # Ajustar el ancho de la columna B (Servei)
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        

        # Obtener los datos de "Serveis oferits" y contar los servicios coincidentes
        

        # Guardar el archivo Excel con el nombre especificado
        wb.save(nombre_archivo)
        self.descargar_btn.show()

    def click_instalar(self):
        # Seleccionar la ubicación donde guardar el archivo
        destino_archivo, _ = QFileDialog.getSaveFileName(
            self, 
            "Guardar archivo como", 
            "", 
            "Archivos Excel (*.xlsx);;Todos los archivos (*)"
        )

        if destino_archivo:
            # Ruta del archivo que deseas descargar
            archivo_fuente = 'excels/facturacio.xlsx'  # Cambia esta ruta a tu archivo

            try:
                # Copiar el archivo a la ubicación seleccionada
                shutil.copy(archivo_fuente, destino_archivo)
                self.descargar_btn.hide()
            except Exception as e:
                self.descargar_btn.hide()
                ventana_error = VentanaError("No s'ha pogut descarregar, informa a Martí", self)
                ventana_error.exec()

class VentanaFidelitat(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Taula de fidelitat")
        self.setMinimumSize(800, 600)
        self.generar_estructura_click()
        self.generar_contenido()
        layoutPrincipal = QGridLayout()
        layoutPrincipal.addWidget(self.ficha_click,0,0)  # Añadir el layout de cuadrícula
        layoutPrincipal.addWidget(self.ficha_botones,0,0)  # Añadir el layout vertical
    
        # Configurar el widget principal y establecer el layout
        contenedor_widget = QWidget(self)
        contenedor_widget.setLayout(layoutPrincipal)
        self.setCentralWidget(contenedor_widget)
        self.parent_window = parent
        self.parent_window.close()
        self.show()

    def generar_contenido(self):
        # Estilos para los botones
        self.ficha_botones = QWidget(self)  
        # Estilo aplicado al widget que contiene el layout

        # Crear el layout principal de la ficha de usuario
        layoutCentral = QGridLayout(self.ficha_botones)

        estilosBoton = """
            QPushButton {
                font-size: 16px;
                border: 1px solid black;
                border-radius: 10px;
                padding: 10px 20px;
            }
            
            QPushButton:hover {
                background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
                border: 2px solid #FFA500;  /* Borde naranja */
                color: white;  /* Cambiar el color del texto a blanco */
            }
        """

        # Crear el layout de cuadrícula
        layoutCentral.setSpacing(20)  # Espaciado entre botones
        layoutCentral.setContentsMargins(50, 50, 50, 50)

        # Crear y configurar los botones
        mesServeis_btn = QPushButton('Clients que han fet més serveis')
        mesServeis_btn.setStyleSheet(estilosBoton)
        mesServeis_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        mesServeis_btn.clicked.connect(self.generar_clients_mes_serveis)
        layoutCentral.addWidget(mesServeis_btn, 0, 0)  # Primera fila, primera columna

        mesAntics_btn = QPushButton('Clients més antics')
        mesAntics_btn.setStyleSheet(estilosBoton)
        mesAntics_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        mesAntics_btn.clicked.connect(self.generar_clients_mes_antics)
        layoutCentral.addWidget(mesAntics_btn, 1, 0)  # Segunda fila, primera columna

        


    def generar_estructura_click(self):
        self.ficha_click = QWidget(self)  

        layoutClick = QGridLayout(self.ficha_click)
        self.generar_tabla()
        self.generar_boton()

        layoutClick.setSpacing(20)  # Espaciado entre botones
        layoutClick.setContentsMargins(50, 50, 50, 50)

        layoutClick.addWidget(self.tablaClientes,0,0)
        layoutClick.addWidget(self.boton,1,0)
        layoutClick.setRowStretch(0,8)
        layoutClick.setRowStretch(1,2)


        self.ficha_click.hide()

    def generar_boton(self):
        self.boton = QPushButton('Tornar')
        self.boton.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                border: 1px solid black;
                border-radius: 10px;
                padding: 10px 20px;
            }
            
            QPushButton:hover {
                background-color: rgba(255, 165, 0, 100);  /* Fondo naranja semi-transparente */
                border: 2px solid #FFA500;  /* Borde naranja */
                color: white;  /* Cambiar el color del texto a blanco */
            }
        """)
        self.boton.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        svg_renderer = QSvgRenderer('assets/back-svg.svg')
        svg_pixmap = QPixmap(64, 64)  # Elige el tamaño que quieras para tu ícono
        svg_pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(svg_pixmap)
        svg_renderer.render(painter)
        painter.end()
        iconCrear = QIcon(svg_pixmap)

        self.boton.setIcon(iconCrear)
        self.boton.setIconSize(svg_pixmap.size())
        self.boton.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.boton.clicked.connect(self.tornar)
    
    def tornar(self):
        self.ficha_botones.show()
        self.ficha_click.hide()

    def generar_tabla(self):
        self.tablaClientes = QTableWidget(self)
        self.limit = 30
        self.offset=0
        self.tablaClientes.setColumnCount(3)  # Número de columnas
        self.tablaClientes.setHorizontalHeaderLabels(["Nom", "Cognoms", "Variable"])  # Nombres de las columnas
        self.tablaClientes.verticalScrollBar().valueChanged.connect(self.load_more_clients)
        self.tablaClientes.setRowCount(0)  # Inicialmente sin filas
        self.configure_table()
        #self.tablaClientes.verticalScrollBar().valueChanged.connect(self.load_more_clients)

    def cambiar_nombre_columna(self, nuevo_nombre):
            # Obtener los nombres actuales de las columnas
        nombres_columnas = []
        for col in range(self.tablaClientes.columnCount()):
            header_item = self.tablaClientes.horizontalHeaderItem(col)
            if header_item:
                nombres_columnas.append(header_item.text())
            else:
                nombres_columnas.append('')  # En caso de que alguna columna no tenga título

        # Cambiar el nombre de la tercera columna (índice 2)
        if len(nombres_columnas) > 2:  # Asegurarse de que haya al menos 3 columnas
            nombres_columnas[2] = nuevo_nombre

        # Aplicar los nuevos nombres a las columnas de la tabla
        self.tablaClientes.setHorizontalHeaderLabels(nombres_columnas)
    def configure_table(self):
        header = self.tablaClientes.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        

        # Ocultar el encabezado vertical
        self.tablaClientes.verticalHeader().setVisible(False)

        # Estilo para la tabla
        self.tablaClientes.setStyleSheet("""
            QTableWidget {
                background: transparent;
                border: 1px solid #d3d3d3;
                border-radius: 10px;
                
            }
            QHeaderView::section {
                background-color: rgba(255, 165, 0, 100);
                border: none;
                font-weight: bold;
                font-size: 18px;
                border-radius: 10px;
            }
            QTableWidget::item {
                background-color: rgba(255, 255, 255, 150);
                border: 1px solid #d3d3d3;
            }

            QTableWidget::item:selected {
                background-color: rgba(255, 255, 255, 150); /* Mantener el fondo igual */
                color: black;  /* El color del texto sigue siendo negro (o cualquier color que prefieras) */
            }
            
        """)

    def generar_clients_mes_serveis(self):
        self.offset = 0
        self.cambiar_nombre_columna("Serveis fets")
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()
        try:
            self.query = '''
                    SELECT 
                    Client.Nom, 
                    Client.Cognoms, 
                    COUNT(Serveis.Id) AS TotalServeis
                    FROM 
                        Client
                    JOIN 
                        Serveis ON Client.Id = Serveis.Client_id
                    GROUP BY 
                        Client.Id
                    HAVING 
                        COUNT(Serveis.Id) > 0
                    ORDER BY 
                        TotalServeis DESC  -- Ordenar de mayor a menor cantidad de servicios
                    LIMIT 
                        ?                 -- Limitar el número de resultados devueltos
                    OFFSET 
                        ?;                 -- Saltar los primeros 0 resultados

                                '''

            params = (f'{self.limit}',f'{self.offset}')

            self.params_list = list(params)
            cursor.execute(self.query, params)

            rows = cursor.fetchall()

            if len(rows) == 0:
                # Si no hay resultados, ocultar la tabla y mostrar un mensaje
                ventana_error = VentanaError("No hi ha cap client que tingui serveis", self)
                ventana_error.exec()

            else:
                # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                self.tablaClientes.setRowCount(0)
                for row in rows:
                    row_position = self.tablaClientes.rowCount()
                    self.tablaClientes.insertRow(row_position)
                    for column, value in enumerate(row):
                        
                        item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                        item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                        self.tablaClientes.setItem(row_position, column, item)

                self.mostrar_click()
                

        finally:
            # Asegurarse de cerrar la conexión
            conn.close()

    def generar_clients_mes_antics(self):
        self.offset = 0
        self.cambiar_nombre_columna("Primer servei")
        conn = sqlite3.connect('data/data.db')
        cursor = conn.cursor()
        try:
            self.query = '''
            SELECT 
                Client.Nom, 
                Client.Cognoms,  
                MIN(Serveis.Fecha) AS FechaAntigua  -- Obtener la fecha más antigua
            FROM 
                Client
            JOIN 
                Serveis ON Client.Id = Serveis.Client_id
            WHERE 
                Client.Client_antic = 1  -- Solo clientes antiguos
            GROUP BY 
                Client.Id  -- Agrupar por cliente
            HAVING 
                COUNT(Serveis.Id) > 0  -- Clientes con al menos un servicio
            ORDER BY 
                FechaAntigua ASC  -- Ordenar por la fecha más antigua
            LIMIT ?  -- Limitar el número de resultados
            OFFSET ?;  -- Saltar los primeros N resultados
        '''

            params = (f'{self.limit}',f'{self.offset}')

            self.params_list = list(params)
            cursor.execute(self.query, params)

            rows = cursor.fetchall()

            if len(rows) == 0:
                # Si no hay resultados, ocultar la tabla y mostrar un mensaje
                ventana_error = VentanaError("No hi ha cap client antic que tingui serveis", self)
                ventana_error.exec()

            else:
                # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                self.tablaClientes.setRowCount(0)
                for row in rows:
                    row_position = self.tablaClientes.rowCount()
                    self.tablaClientes.insertRow(row_position)
                    for column, value in enumerate(row):
                        if column == 2:
                            value = self.fecha_a_string(value)
                        item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                        item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                        self.tablaClientes.setItem(row_position, column, item)

                self.mostrar_click()
                

        finally:
            # Asegurarse de cerrar la conexión
            conn.close()

    def load_more_clients(self):
        if self.tablaClientes.verticalScrollBar().value() == self.tablaClientes.verticalScrollBar().maximum():

            self.offset += self.limit
            conn = sqlite3.connect('data/data.db')
            cursor = conn.cursor()
            self.params_list[-1] = f'{self.offset}'
            params = tuple(self.params_list)
            try:
                cursor.execute(self.query, params)

                rows = cursor.fetchall()

                if len(rows) == 0:
                    pass
                else:
                    # Si hay resultados, mostrar la tabla y ocultar el mensaje de "sin datos"
                    for row in rows:
                        row_position = self.tablaClientes.rowCount()
                        self.tablaClientes.insertRow(row_position)
                        for column, value in enumerate(row):
                            item = QTableWidgetItem(str(value))  # Convertir el valor a cadena
                            item.setFont(QFont("Arial", 16))  # Ajustar el tamaño de la fuente
                            self.tablaClientes.setItem(row_position, column, item)

                
            finally:
                # Asegurarse de cerrar la conexión
                conn.close()

    def fecha_a_string(self,fecha_str):
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')  # Convierte la cadena a objeto datetime
        return fecha_obj.strftime('%d/%m/%Y')

    def mostrar_click(self):
        self.ficha_botones.hide()
        self.ficha_click.show()


    
    


      







if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec())
